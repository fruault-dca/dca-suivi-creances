"""
Suivi des Créances Clients - Application Streamlit
Backend : Google Sheets (base de données partagée)
"""
import streamlit as st
import pandas as pd
import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import gspread
from gspread.exceptions import APIError
from google.oauth2.service_account import Credentials
import time, random

st.set_page_config(page_title="Suivi Créances Clients", page_icon="📊", layout="wide")

# Charte graphique DCA (Design Constructions et Associés)
st.markdown("""
<style>
/* Police Segoe UI sur toute l'app */
html, body, [class*="css"], .stMarkdown, .stButton, .stTextInput, .stDataFrame {
    font-family: "Segoe UI", -apple-system, BlinkMacSystemFont, sans-serif;
}

/* Titres en bleu marine DCA */
h1, h2, h3, h4 { color: #2C3E50; font-weight: 600; }

/* Métriques : valeur en bleu marine, label en gris */
[data-testid="stMetricValue"] { color: #2C3E50; font-weight: 700; }
[data-testid="stMetricLabel"] { color: #546E7A; }

/* Boutons primaires en vert DCA */
.stButton > button[kind="primary"] {
    background-color: #60A020;
    border-color: #4A7A18;
}
.stButton > button[kind="primary"]:hover {
    background-color: #4A7A18;
    border-color: #355A10;
}

/* Sidebar plus contrastée */
[data-testid="stSidebar"] { background-color: #F0F0F0; }

/* Headers de tableau en bleu marine */
.stDataFrame thead tr th {
    background-color: #2C3E50 !important;
    color: white !important;
    font-weight: 600;
}
</style>
""", unsafe_allow_html=True)

# ============================================================
# CONFIGURATION GOOGLE SHEETS
# ============================================================
SCOPES = [
    'https://www.googleapis.com/auth/spreadsheets',
    'https://www.googleapis.com/auth/drive',
]

HEADERS = {
    'creances': ['id', 'comp_aux_num', 'comp_aux_lib', 'piece_ref', 'piece_date',
                 'ecriture_date', 'journal_code', 'ecriture_lib', 'debit', 'credit',
                 'ecriture_let', 'import_date'],
    'dossiers': ['ref_client', 'code_affaire', 'client', 'email1', 'email2',
                 'type_projet', 'adresse', 'cp', 'ville', 'constructeur',
                 'agence', 'commercial', 'conducteur', 'etat', 'stade', 'type_contrat',
                 'contrat_ht', 'contrat_ttc', 'contrat_rev_ht', 'contrat_rev_ttc',
                 'avenants_ht', 'avenants_ttc', 'date_signature', 'date_reception'],
    'mapping': ['piece_ref', 'ref_client', 'comp_aux_num',
                'date_facture', 'situation'],
    'notes': ['id', 'ref_client', 'comp_aux_num', 'date_note', 'auteur', 'note',
              'action', 'echeance', 'statut', 'assigne_a'],
    'users': ['email', 'nom_affichage', 'actif', 'password_hash', 'role'],
    'contentieux': ['ref_client', 'comp_aux_num', 'responsable',
                    'date_passage', 'commentaire',
                    'provision_risque', 'provision_creances_douteuses'],
    'resumes': ['comp_aux_num', 'ref_client', 'resume',
                'action_resume', 'responsable_action',
                'date_recouvrement', 'nature_creance',
                'date_maj', 'auteur'],
    'consignations': ['comp_aux_num', 'ref_client', 'montant_consigne',
                      'date_consignation', 'commentaire'],
}

# Liste des natures de créance (export Direction)
NATURES_CREANCE = [
    '—',
    'Procédure judiciaire',
    'Avoir à émettre',
    'A suivre par Eric',
    'En cours',
    'Relance huissier',
    'Travaux à réaliser',
    'Travaux terminés à relancer',
    'Facturation sous-traitant',
]


@st.cache_resource
def get_gspread_client():
    """Se connecte à Google Sheets via le compte de service."""
    creds_dict = dict(st.secrets["gcp_service_account"])
    creds = Credentials.from_service_account_info(creds_dict, scopes=SCOPES)
    return gspread.authorize(creds)


@st.cache_resource
def get_spreadsheet():
    client = get_gspread_client()
    return client.open_by_key(st.secrets["google"]["sheet_id"])


@st.cache_resource
def get_ws(name):
    ss = get_spreadsheet()
    try:
        return ss.worksheet(name)
    except gspread.exceptions.WorksheetNotFound:
        # Crée la feuille manquante à la volée
        headers = HEADERS.get(name, [])
        ws = _with_retry(ss.add_worksheet, title=name, rows=100,
                         cols=max(len(headers), 1))
        if headers:
            _with_retry(ws.update, values=[headers], range_name='A1')
        return ws


def _with_retry(fn, *args, **kwargs):
    """Appelle fn avec retry exponentiel sur les erreurs 429 (quota)."""
    for attempt in range(5):
        try:
            return fn(*args, **kwargs)
        except APIError as e:
            code = getattr(e.response, 'status_code', None) if hasattr(e, 'response') else None
            msg = str(e)
            if code == 429 or '429' in msg or 'Quota exceeded' in msg:
                wait = (2 ** attempt) + random.random()
                time.sleep(wait)
                continue
            raise
    raise APIError({"error": {"message": "Quota Google Sheets dépassé après 5 tentatives"}})


@st.cache_resource
def ensure_headers():
    """Écrit les en-têtes dans chaque onglet si absentes. Une seule fois par session."""
    ss = get_spreadsheet()
    existing_sheets = {s.title for s in _with_retry(ss.worksheets)}
    for sheet_name, headers in HEADERS.items():
        if sheet_name not in existing_sheets:
            _with_retry(ss.add_worksheet, title=sheet_name, rows=100, cols=len(headers))
        ws = ss.worksheet(sheet_name)
        first_row = _with_retry(ws.row_values, 1)
        if first_row != headers:
            _with_retry(ws.update, values=[headers], range_name='A1')
    return True


@st.cache_data(ttl=600, show_spinner=False)
def read_sheet(name):
    """Lit un onglet et retourne un DataFrame (cache 10 min pour économiser le quota)."""
    ws = get_ws(name)
    records = _with_retry(ws.get_all_records)
    df = pd.DataFrame(records)
    if df.empty:
        df = pd.DataFrame(columns=HEADERS[name])
    return df


def clear_cache():
    st.cache_data.clear()


def replace_sheet(name, df):
    """Écrase complètement un onglet avec un DataFrame."""
    ws = get_ws(name)
    _with_retry(ws.clear)
    headers = HEADERS[name]
    if df.empty:
        _with_retry(ws.update, values=[headers], range_name='A1')
    else:
        df2 = df.copy()
        for h in headers:
            if h not in df2.columns:
                df2[h] = ''
        df2 = df2[headers].fillna('').astype(str)
        values = [headers] + df2.values.tolist()
        _with_retry(ws.update, values=values, range_name='A1')
    clear_cache()


def append_row(name, row_dict):
    ws = get_ws(name)
    headers = HEADERS[name]
    row = [str(row_dict.get(h, '')) for h in headers]
    _with_retry(ws.append_row, row, value_input_option='USER_ENTERED')
    clear_cache()


def update_cell_by_id(name, row_id, column, new_value):
    """Met à jour une cellule en trouvant la ligne par son id."""
    ws = get_ws(name)
    headers = HEADERS[name]
    col_idx = headers.index(column) + 1
    id_col_idx = headers.index('id') + 1
    cell = ws.find(str(row_id), in_column=id_col_idx)
    if cell:
        ws.update_cell(cell.row, col_idx, new_value)
        clear_cache()


def update_row_by_id(name, row_id, updates: dict):
    """Met à jour plusieurs colonnes d'une ligne en une seule passe."""
    ws = get_ws(name)
    headers = HEADERS[name]
    id_col_idx = headers.index('id') + 1
    cell = _with_retry(ws.find, str(row_id), in_column=id_col_idx)
    if not cell:
        return
    cells = []
    for col, val in updates.items():
        if col in headers:
            c = ws.cell(cell.row, headers.index(col) + 1)
            c.value = val
            cells.append(c)
    if cells:
        _with_retry(ws.update_cells, cells)
        clear_cache()


def delete_row_by_id(name, row_id):
    ws = get_ws(name)
    headers = HEADERS[name]
    id_col_idx = headers.index('id') + 1
    cell = _with_retry(ws.find, str(row_id), in_column=id_col_idx)
    if cell:
        _with_retry(ws.delete_rows, cell.row)
        clear_cache()
        clear_cache()


def next_id(df):
    if df.empty or 'id' not in df.columns:
        return 1
    try:
        return int(pd.to_numeric(df['id'], errors='coerce').max()) + 1
    except (ValueError, TypeError):
        return 1


# ============================================================
# VÉRIFICATION DE LA CONFIG
# ============================================================
def hash_pwd(plain: str) -> str:
    """Hash sécurisé d'un mot de passe (PBKDF2-SHA256 + salt aléatoire)."""
    import hashlib, secrets as _secrets
    salt = _secrets.token_hex(16)
    h = hashlib.pbkdf2_hmac('sha256', plain.encode('utf-8'),
                             salt.encode('utf-8'), 100_000)
    return f"pbkdf2$100000${salt}${h.hex()}"


def check_pwd(plain: str, stored: str) -> bool:
    """Vérifie un mot de passe contre son hash stocké."""
    import hashlib
    if not stored or '$' not in stored:
        return False
    try:
        algo, iters, salt, hexhash = stored.split('$')
        if algo != 'pbkdf2':
            return False
        h = hashlib.pbkdf2_hmac('sha256', plain.encode('utf-8'),
                                 salt.encode('utf-8'), int(iters))
        return h.hex() == hexhash
    except Exception:
        return False


def login_user(email: str, nom: str):
    st.session_state['auth_email'] = email
    st.session_state['auth_nom'] = nom


def logout_user():
    for k in ('auth_email', 'auth_nom', 'manual_user'):
        if k in st.session_state:
            del st.session_state[k]


def is_logged_in() -> bool:
    return 'auth_email' in st.session_state


def show_login():
    """Écran de connexion par mot de passe."""
    st.markdown(
        "<style>.block-container { max-width: 480px; padding-top: 4rem; }</style>",
        unsafe_allow_html=True,
    )
    st.title("💼 Suivi Créances DCA")
    st.caption("Connectez-vous pour accéder à l'application")

    df_users = read_sheet('users')
    if df_users.empty:
        st.error("Aucun utilisateur n'est encore enregistré dans le système.")
        st.info("L'administrateur doit d'abord créer des utilisateurs dans "
                "l'onglet Import → Utilisateurs.")
        st.stop()

    # Crée les colonnes manquantes pour éviter les erreurs
    for col, default in [('email', ''), ('nom_affichage', ''),
                          ('actif', 'oui'), ('password_hash', ''),
                          ('role', 'user')]:
        if col not in df_users.columns:
            df_users[col] = default
        df_users[col] = df_users[col].fillna(default).astype(str)

    df_users['email'] = df_users['email'].str.lower()
    df_users['actif'] = df_users['actif'].str.lower()
    df_users = df_users[df_users['actif'].isin(
        ('oui', 'true', '1', 'yes', ''))]

    with st.form("login_form"):
        email = st.text_input("Email").strip().lower()
        password = st.text_input("Mot de passe", type="password")
        col_a, col_b = st.columns(2)
        submitted = col_a.form_submit_button("Se connecter", type="primary",
                                              use_container_width=True)
        signup_mode = col_b.form_submit_button(
            "1ère connexion (définir mot de passe)",
            use_container_width=True)

    if not (submitted or signup_mode):
        st.stop()

    if not email or '@' not in email:
        st.error("Email invalide.")
        st.stop()

    user_row = df_users[df_users['email'] == email]
    if user_row.empty:
        st.error("Aucun compte actif ne correspond à cet email. "
                 "Contactez l'administrateur.")
        st.stop()

    user = user_row.iloc[0]
    nom = user.get('nom_affichage', email) or email
    stored_hash = str(user.get('password_hash', '') or '')

    if signup_mode:
        # 1ère connexion : autorisé seulement si pas encore de mdp défini
        if stored_hash:
            st.error("Un mot de passe est déjà défini pour ce compte. "
                     "Utilisez 'Se connecter' (ou demandez une réinitialisation).")
            st.stop()
        if len(password) < 6:
            st.error("Le mot de passe doit faire au moins 6 caractères.")
            st.stop()
        new_hash = hash_pwd(password)
        # Met à jour la ligne dans le sheet
        df_full = read_sheet('users')
        df_full.loc[df_full['email'].astype(str).str.lower() == email,
                    'password_hash'] = new_hash
        replace_sheet('users', df_full)
        login_user(email, nom)
        st.success(f"✅ Mot de passe défini. Bienvenue {nom}.")
        st.rerun()

    # Connexion normale
    if not stored_hash:
        st.warning("Ce compte n'a pas encore de mot de passe défini. "
                   "Cliquez sur '1ère connexion' pour en créer un.")
        st.stop()
    if not check_pwd(password, stored_hash):
        st.error("Mot de passe incorrect.")
        st.stop()
    login_user(email, nom)
    st.rerun()


def current_user():
    """Retourne l'utilisateur courant {email, nom} si authentifié."""
    if 'auth_email' in st.session_state:
        return {
            'email': st.session_state['auth_email'],
            'nom': st.session_state.get('auth_nom',
                                          st.session_state['auth_email']),
        }
    # Fallback dev/local
    if 'manual_user' in st.session_state:
        return st.session_state['manual_user']
    return None


def get_active_users():
    """Liste des utilisateurs actifs (depuis la feuille users)."""
    df_users = read_sheet('users')
    if df_users.empty:
        return []
    df_users = df_users[df_users.get('actif', 'oui').astype(str).str.lower()
                          .isin(['oui', 'true', '1', 'yes', ''])]
    return [{'email': r['email'],
             'nom': r.get('nom_affichage', r['email']) or r['email']}
            for _, r in df_users.iterrows() if r.get('email')]


def check_config():
    try:
        if "gcp_service_account" not in st.secrets:
            return False, "Secret `gcp_service_account` manquant"
        if "google" not in st.secrets or "sheet_id" not in st.secrets["google"]:
            return False, "Secret `google.sheet_id` manquant"
        ensure_headers()
        return True, "OK"
    except Exception as e:
        return False, f"Erreur connexion Google Sheets : {e}"


# ============================================================
# HELPERS PARSING
# ============================================================
def to_float(val):
    if pd.isna(val) or val is None:
        return 0.0
    try:
        return float(str(val).replace(',', '.').replace(' ', '').replace('\xa0', ''))
    except (ValueError, TypeError):
        return 0.0


def to_str(val):
    if pd.isna(val) or val is None or str(val).lower() == 'nan':
        return ''
    # float entier → int (évite "830.0" pour un code dossier)
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    return str(val).strip()


def fr_date(s):
    """Convertit une date ISO YYYY-MM-DD (ou autre) en JJ/MM/AAAA pour affichage."""
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return ''
    s = str(s).strip()
    if not s or s.lower() == 'nan':
        return ''
    # Déjà au format français → on garde
    if len(s) >= 10 and s[2] == '/' and s[5] == '/':
        return s[:10]
    try:
        dt = pd.to_datetime(s, errors='coerce', dayfirst=False)
        if pd.isna(dt):
            dt = pd.to_datetime(s, errors='coerce', dayfirst=True)
        if pd.isna(dt):
            return s
        return dt.strftime('%d/%m/%Y')
    except Exception:
        return s


def fr_series(s):
    """Convertit une Series de dates en strings JJ/MM/AAAA."""
    return s.fillna('').astype(str).apply(fr_date)


def to_date_obj(s):
    """Convertit une string en objet date Python (pour Excel reconnaissable)."""
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return None
    s = str(s).strip()
    if not s or s.lower() == 'nan':
        return None
    try:
        dt = pd.to_datetime(s, errors='coerce', dayfirst=False)
        if pd.isna(dt):
            dt = pd.to_datetime(s, errors='coerce', dayfirst=True)
        if pd.isna(dt):
            return None
        return dt.date()
    except Exception:
        return None


def format_date_fec(d):
    s = to_str(d).strip()
    if not s:
        return ''
    # Format FEC standard YYYYMMDD
    if len(s) == 8 and s.isdigit():
        return f"{s[:4]}-{s[4:6]}-{s[6:]}"
    # Format ISO déjà bon
    if len(s) >= 10 and s[4] == '-' and s[7] == '-':
        return s[:10]
    # Format DD/MM/YYYY ou DD-MM-YYYY → on bascule en YYYY-MM-DD
    try:
        dt = pd.to_datetime(s, errors='coerce', dayfirst=True)
        if pd.notna(dt):
            return dt.date().isoformat()
    except Exception:
        pass
    return s


def parse_fec(file_content):
    cols = ['JournalCode', 'JournalLib', 'EcritureNum', 'EcritureDate', 'CompteNum',
            'CompteLib', 'CompAuxNum', 'CompAuxLib', 'PieceRef', 'PieceDate',
            'EcritureLib', 'Debit', 'Credit', 'EcritureLet', 'DateLet', 'ValidDate',
            'Montantdevise', 'Idevise']
    for enc in ['utf-8', 'latin-1', 'cp1252']:
        try:
            df = pd.read_csv(io.BytesIO(file_content), sep='\t', encoding=enc,
                             skiprows=1, header=None, names=cols, dtype=str)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("Impossible de décoder le FEC")

    # 411xxx : créances clients normales / 416xxx : clients douteux ou litigieux
    mask = df['CompteNum'].str.startswith('411', na=False) \
         | df['CompteNum'].str.startswith('416', na=False)
    clients = df[mask].copy()
    clients['Debit'] = clients['Debit'].apply(to_float)
    clients['Credit'] = clients['Credit'].apply(to_float)
    return clients


def parse_crm(file_content, sheet_name='Liste complète'):
    df = pd.read_excel(io.BytesIO(file_content), sheet_name=sheet_name,
                       header=None, dtype=str)
    headers = df.iloc[1].tolist()
    data = df.iloc[2:].copy()
    data.columns = headers
    data = data[data.iloc[:, 0].astype(str).str.strip() != 'Totaux :']
    data = data[data.iloc[:, 0].notna()]
    return data


def load_creances_enrichies(only_open=True):
    df_c = read_sheet('creances')
    df_m = read_sheet('mapping')
    df_d = read_sheet('dossiers')

    if df_c.empty:
        return pd.DataFrame()

    df_c['debit'] = pd.to_numeric(df_c['debit'], errors='coerce').fillna(0)
    df_c['credit'] = pd.to_numeric(df_c['credit'], errors='coerce').fillna(0)
    df_c['solde'] = df_c['debit'] - df_c['credit']

    if only_open:
        df_c = df_c[(df_c['ecriture_let'].isna()) | (df_c['ecriture_let'] == '')]

        # Auto-rapprochement FIFO par client :
        # Pour chaque client, on prend tous ses encaissements non lettrés (crédits)
        # et on les impute sur ses plus anciennes factures (débits) dans l'ordre.
        # Ainsi, même si l'encaissement n'a pas la même piece_ref que la facture,
        # le rapprochement se fait correctement au niveau du compte client.
        if not df_c.empty:
            kept_rows = []
            # Tri sur date pour FIFO (plus anciennes d'abord)
            df_c['_date_sort'] = pd.to_datetime(df_c['ecriture_date'],
                                                 errors='coerce', dayfirst=True)

            for comp_num, grp in df_c.groupby('comp_aux_num', dropna=False):
                debits = grp[grp['debit'] > 0].sort_values('_date_sort').copy()
                credits_sum = grp['credit'].sum()

                # Impute les crédits sur les débits les plus anciens (FIFO)
                remaining = credits_sum
                new_soldes = []
                for _, d in debits.iterrows():
                    montant = d['debit']
                    if remaining >= montant - 0.01:
                        remaining -= montant
                        new_soldes.append(0.0)  # facture totalement soldée
                    elif remaining > 0:
                        new_soldes.append(montant - remaining)
                        remaining = 0
                    else:
                        new_soldes.append(montant)

                debits = debits.assign(solde=new_soldes)
                # On retire les factures soldées
                debits = debits[debits['solde'].abs() > 0.01]
                kept_rows.append(debits)

                # Si après avoir imputé tous les débits il reste du crédit (avoir/trop-percu),
                # on le remonte comme ligne négative (rare)
                if remaining < -0.01:
                    dummy = grp[grp['credit'] > 0].sort_values('_date_sort').head(1).copy()
                    if not dummy.empty:
                        dummy = dummy.assign(solde=-abs(remaining), debit=0, credit=abs(remaining))
                        kept_rows.append(dummy)

            df_c = pd.concat(kept_rows, ignore_index=True) if kept_rows else \
                   df_c.iloc[0:0]
            if '_date_sort' in df_c.columns:
                df_c = df_c.drop(columns=['_date_sort'])

    if not df_m.empty and 'piece_ref' in df_m.columns:
        # Normalise piece_ref des deux côtés (enlève zéros de tête sur chaque segment)
        # pour matcher FEC "22/1" avec PROGEMI "22/0000001"
        def norm_piece(s):
            s = str(s).strip()
            if not s:
                return ''
            parts = s.split('/')
            return '/'.join(p.lstrip('0') or '0' for p in parts)

        df_c['_pk'] = df_c['piece_ref'].apply(norm_piece)
        map_cols = ['piece_ref', 'ref_client']
        for opt in ('date_facture', 'situation'):
            if opt in df_m.columns:
                map_cols.append(opt)
        df_m2 = df_m[map_cols].copy()
        df_m2['_pk'] = df_m2['piece_ref'].apply(norm_piece)
        # Priorité aux refs réelles sur "__HORS_CRM__" en cas de doublon de _pk
        df_m2['_priority'] = (df_m2['ref_client'] != '__HORS_CRM__').astype(int)
        df_m2 = df_m2.sort_values('_priority', ascending=False) \
            .drop_duplicates('_pk')
        keep_cols = ['_pk', 'ref_client']
        for opt in ('date_facture', 'situation'):
            if opt in df_m2.columns:
                keep_cols.append(opt)
        df_m2 = df_m2[keep_cols]
        df_c = df_c.merge(df_m2, on='_pk', how='left').drop(columns=['_pk'])
    else:
        df_c['ref_client'] = ''

    # Propagation : si un client a au moins une facture avec une vraie ref CRM,
    # toutes ses autres factures Hors CRM héritent de cette ref
    # (cas typique : client en contentieux dont seules certaines factures sont
    # passées en PROGEMI, les autres ont été marquées Hors CRM par erreur)
    if 'ref_client' in df_c.columns:
        mask_hors = df_c['ref_client'] == '__HORS_CRM__'
        real_refs = df_c[
            ~mask_hors
            & df_c['ref_client'].fillna('').astype(bool)
            & (df_c['ref_client'] != '__HORS_CRM__')
        ].drop_duplicates('comp_aux_num')[['comp_aux_num', 'ref_client']] \
         .rename(columns={'ref_client': '_ref_real'})
        if not real_refs.empty:
            df_c = df_c.merge(real_refs, on='comp_aux_num', how='left')
            propagate = mask_hors & df_c['_ref_real'].fillna('').astype(bool)
            df_c.loc[propagate, 'ref_client'] = df_c.loc[propagate, '_ref_real']
            df_c = df_c.drop(columns=['_ref_real'])

    if not df_d.empty:
        dos_cols = ['ref_client', 'client', 'commercial', 'conducteur', 'agence', 'etat',
                    'stade', 'contrat_ttc', 'date_reception']
        df_d_small = df_d[[c for c in dos_cols if c in df_d.columns]].copy()

        # Match tolérant aux zéros de tête : "549" matche "00549"
        # On crée une clé normalisée des deux côtés et on fait le merge dessus.
        import re as _re_d
        def _norm_ref(s):
            s = str(s or '').strip()
            if not s:
                return ''
            # Pour les refs groupées CRM "830/831", on prend la version brute (sans zéros)
            parts = _re_d.split(r'[/,;]+', s)
            return '/'.join(p.strip().lstrip('0') or '0' for p in parts if p.strip())

        df_c['_rk'] = df_c['ref_client'].apply(_norm_ref)
        df_d_small['_rk'] = df_d_small['ref_client'].apply(_norm_ref)
        # Conserve la ref CRM d'origine, ignore le doublon ref_client de df_d_small
        df_d_small = df_d_small.drop(columns=['ref_client']) \
            .drop_duplicates('_rk')
        df_c = df_c.merge(df_d_small, on='_rk', how='left').drop(columns=['_rk'])
    else:
        for col in ['client', 'commercial', 'conducteur', 'agence', 'etat', 'stade',
                    'contrat_ttc', 'date_reception']:
            df_c[col] = ''

    # Pour les vrais Hors CRM (pas de CRM connue), affichage "Hors CRM"
    mask_hors_final = df_c['ref_client'] == '__HORS_CRM__'
    df_c.loc[mask_hors_final, 'ref_client'] = 'Hors CRM'
    df_c.loc[mask_hors_final, 'client'] = df_c.loc[mask_hors_final, 'comp_aux_lib']

    # Jours de retard : priorité à la date de facture du PROGEMI (date_facture),
    # puis piece_date du FEC, puis fallback ecriture_date
    today = pd.Timestamp(datetime.now().date())
    if 'date_facture' in df_c.columns:
        df_c['_dt_progemi'] = pd.to_datetime(df_c['date_facture'],
                                              errors='coerce', dayfirst=True)
    else:
        df_c['_dt_progemi'] = pd.NaT
    if 'piece_date' in df_c.columns:
        df_c['_dt_piece'] = pd.to_datetime(df_c['piece_date'],
                                            errors='coerce', dayfirst=True)
    else:
        df_c['_dt_piece'] = pd.NaT
    df_c['_dt_ecr'] = pd.to_datetime(df_c['ecriture_date'],
                                      errors='coerce', dayfirst=True)
    df_c['_dt'] = df_c['_dt_progemi'].fillna(df_c['_dt_piece']).fillna(df_c['_dt_ecr'])
    df_c['jours_retard'] = (today - df_c['_dt']).dt.days
    df_c['jours_retard'] = df_c['jours_retard'].fillna(0).astype(int).clip(lower=0)
    # Date de facture effective (priorité PROGEMI > FEC piece_date > ecriture_date)
    df_c['date_facture_eff'] = df_c['_dt'].dt.strftime('%Y-%m-%d')
    # Filet de sécurité : si rien n'a pu être daté, garder l'écriture brute
    df_c['date_facture_eff'] = df_c['date_facture_eff'].fillna(
        df_c['ecriture_date'].fillna('').astype(str))
    df_c['date_facture_eff'] = df_c['date_facture_eff'].fillna('')
    df_c = df_c.drop(columns=['_dt', '_dt_progemi', '_dt_piece', '_dt_ecr'])

    # Flag chantier livré : si date_reception est renseignée
    if 'date_reception' in df_c.columns:
        # Nettoie la date de réception : "2025-08-01 00:00:00" -> "2025-08-01"
        df_c['date_reception'] = df_c['date_reception'].fillna('').astype(str) \
            .str.split(' ').str[0].str.strip()
        df_c['est_livre'] = df_c['date_reception'] != ''
    else:
        df_c['est_livre'] = False

    # Merge consignations huissier (montant consigné par client)
    df_cons = read_sheet('consignations')
    if not df_cons.empty and 'comp_aux_num' in df_cons.columns:
        df_cons_small = df_cons[['comp_aux_num', 'montant_consigne']].copy()
        df_cons_small['montant_consigne'] = pd.to_numeric(
            df_cons_small['montant_consigne'], errors='coerce').fillna(0)
        df_cons_small = df_cons_small.drop_duplicates('comp_aux_num')
        df_c = df_c.merge(df_cons_small, on='comp_aux_num', how='left')
        df_c['montant_consigne'] = df_c['montant_consigne'].fillna(0)
    else:
        df_c['montant_consigne'] = 0.0

    # Flag contentieux + responsable + provisions
    # Match sur ref_client (cas dossier CRM) OU comp_aux_num (cas client FEC sans CRM)
    df_ct = read_sheet('contentieux')
    df_c['contentieux'] = False
    df_c['responsable'] = ''
    df_c['provision_risque'] = 0.0
    df_c['provision_creances_douteuses'] = 0.0

    if not df_ct.empty and 'ref_client' in df_ct.columns:
        prov_cols = [c for c in ['provision_risque', 'provision_creances_douteuses']
                     if c in df_ct.columns]

        for _, ct in df_ct.iterrows():
            ref = str(ct.get('ref_client', '') or '').strip()
            comp = str(ct.get('comp_aux_num', '') or '').strip()
            resp = str(ct.get('responsable', '') or '').strip()
            if not resp:
                continue

            mask = pd.Series(False, index=df_c.index)
            if ref:
                mask |= (df_c['ref_client'].astype(str) == ref)
            if comp:
                mask |= (df_c['comp_aux_num'].astype(str) == comp)

            if mask.any():
                df_c.loc[mask, 'contentieux'] = True
                df_c.loc[mask, 'responsable'] = resp
                for pc in prov_cols:
                    val = pd.to_numeric(ct.get(pc, 0), errors='coerce')
                    if pd.isna(val):
                        val = 0.0
                    df_c.loc[mask, pc] = val

    return df_c.sort_values('solde', ascending=False)


# ============================================================
# PAGES
# ============================================================
def page_import():
    st.header("📥 Import des données")

    tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
        "FEC", "CRM (Chantiers)", "Mapping factures",
        "Contentieux", "Consignations huissier", "Utilisateurs"])

    with tab1:
        st.markdown("**Import du Fichier d'Écritures Comptables**")
        st.caption("Extrait les comptes 411xxx (créances clients) et 416xxx (clients douteux/litigieux). "
                   "Les écritures lettrées sont marquées comme soldées.")
        fec_file = st.file_uploader("Fichier FEC (.txt)", type=['txt'], key='fec')
        if fec_file and st.button("Importer le FEC", type="primary"):
            try:
                with st.spinner("Analyse du FEC et écriture dans Google Sheets..."):
                    clients = parse_fec(fec_file.read())
                    rows = []
                    for i, (_, r) in enumerate(clients.iterrows(), 1):
                        rows.append({
                            'id': i,
                            'comp_aux_num': r['CompAuxNum'],
                            'comp_aux_lib': r['CompAuxLib'],
                            'piece_ref': r['PieceRef'],
                            'piece_date': format_date_fec(r['PieceDate']),
                            'ecriture_date': format_date_fec(r['EcritureDate']),
                            'journal_code': r['JournalCode'],
                            'ecriture_lib': r['EcritureLib'],
                            'debit': r['Debit'],
                            'credit': r['Credit'],
                            'ecriture_let': to_str(r['EcritureLet']),
                            'import_date': datetime.now().isoformat(),
                        })
                    df_new = pd.DataFrame(rows)
                    replace_sheet('creances', df_new)
                st.success(f"✅ {len(clients)} écritures clients importées")
            except Exception as e:
                st.error(f"Erreur : {e}")

    with tab2:
        st.markdown("**Import de l'export CRM (Chantiers)**")
        crm_file = st.file_uploader("Export CRM (.xlsx)", type=['xlsx'], key='crm')
        if crm_file:
            try:
                xl = pd.ExcelFile(io.BytesIO(crm_file.getvalue()))
                default_idx = xl.sheet_names.index('Liste complète') \
                    if 'Liste complète' in xl.sheet_names else 0
                sheet = st.selectbox("Feuille à importer", xl.sheet_names, index=default_idx)
                if st.button("Importer le CRM", type="primary"):
                    with st.spinner("Analyse du CRM et écriture dans Google Sheets..."):
                        data = parse_crm(crm_file.getvalue(), sheet_name=sheet)

                        # Lookup tolérant aux accents/casse/espaces
                        import unicodedata as _ud, re as _re3
                        def _norm_col(s):
                            s = str(s).strip().lower()
                            s = _ud.normalize('NFKD', s).encode('ascii', 'ignore').decode()
                            return _re3.sub(r'[^a-z0-9]+', '', s)
                        _col_lookup = {_norm_col(c): c for c in data.columns}

                        def g(row, col):
                            key = _norm_col(col)
                            actual = _col_lookup.get(key)
                            if actual is None:
                                return ''
                            return to_str(row.get(actual, ''))

                        def gf(row, col):
                            key = _norm_col(col)
                            actual = _col_lookup.get(key)
                            if actual is None:
                                return ''
                            v = row.get(actual, '')
                            if pd.isna(v) or str(v).lower() == 'nan' or str(v).strip() == '':
                                return ''
                            return to_float(v)

                        # Debug : affiche les colonnes détectées
                        with st.expander("🔍 Colonnes détectées dans le fichier CRM"):
                            st.write(list(data.columns))

                        rows = []
                        for _, r in data.iterrows():
                            ref = g(r, 'Ref client')
                            if not ref:
                                continue
                            rows.append({
                                'ref_client': ref,
                                'code_affaire': g(r, 'N°Compta/Code Affaire'),
                                'client': g(r, 'Client(s)'),
                                'email1': g(r, 'Client Email 1'),
                                'email2': g(r, 'Client Email 2'),
                                'type_projet': g(r, 'Type de projet'),
                                'adresse': g(r, 'Adresse du projet'),
                                'cp': g(r, 'CP'),
                                'ville': g(r, 'Ville'),
                                'constructeur': g(r, 'Constructeur'),
                                'agence': g(r, 'Agence'),
                                'commercial': g(r, 'Commercial'),
                                'conducteur': g(r, 'Conducteur de travaux'),
                                'etat': g(r, 'Etat'),
                                'stade': g(r, "Stade d'avancement"),
                                'type_contrat': g(r, 'Type de contrat'),
                                'contrat_ht': gf(r, 'Contrat HT'),
                                'contrat_ttc': gf(r, 'Contrat TTC'),
                                'contrat_rev_ht': gf(r, 'Contrat révisé HT'),
                                'contrat_rev_ttc': gf(r, 'Contrat révisé TTC'),
                                'avenants_ht': gf(r, 'Avenants HT'),
                                'avenants_ttc': gf(r, 'Avenants TTC'),
                                'date_signature': g(r, 'Date de signature du contrat'),
                                'date_reception': g(r, 'Date de réception'),
                            })
                        replace_sheet('dossiers', pd.DataFrame(rows))
                    st.success(f"✅ {len(rows)} dossiers importés")
            except Exception as e:
                st.error(f"Erreur : {e}")

    with tab3:
        st.markdown("**Import du fichier de facturation** (lien facture ↔ dossier CRM)")
        st.info(
            "Fichier CSV/Excel contenant les correspondances entre numéros de facture "
            "(présents dans le FEC) et références dossier (présentes dans le CRM).\n\n"
            "**Colonnes attendues** :\n"
            "- `piece_ref` — numéro de facture (ex: `26/0000002`, `21/52`)\n"
            "- `ref_client` — référence dossier CRM (ex: `00655`, `976`)\n"
            "- `comp_aux_num` — *optionnel*, code comptable client (ex: `CROBERT`)\n\n"
            "💡 Les noms de colonnes peuvent varier : `numero_facture`, `num_facture`, "
            "`facture`, `ref_dossier`, `dossier` sont aussi reconnus."
        )

        map_file = st.file_uploader("Fichier de facturation (.csv ou .xlsx)",
                                    type=['csv', 'xlsx'], key='map')

        # Option pour cumuler (ajouter) ou remplacer
        mode = st.radio("Mode d'import",
                        ["Remplacer tout le mapping existant",
                         "Ajouter / mettre à jour (cumul multi-années)"],
                        horizontal=True, key='map_mode')

        if map_file and st.button("Importer le fichier de facturation", type="primary"):
            try:
                # Aliases (normalisés sans accent ni espace)
                aliases_piece = ['piece_ref', 'numero_facture', 'num_facture',
                                 'facture', 'n_facture', 'no_facture', 'piece',
                                 'nfacture', 'numfacture']
                aliases_ref = ['ref_client', 'ref_dossier', 'dossier', 'ref',
                               'code_affaire', 'n_compta', 'code_dossier',
                               'codedossier', 'refclient', 'refdossier']
                aliases_comp = ['comp_aux_num', 'code_client', 'code_compta',
                                'code_comptable']
                aliases_date = ['date', 'date_facture', 'datefacture',
                                'date_fact', 'date_piece', 'date_emission']
                aliases_situation = ['situation', 'etat', 'avancement',
                                      'stade', 'situation_facture',
                                      'etat_avancement']

                def normalize(s):
                    import unicodedata, re
                    s = str(s).strip().lower()
                    s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode()
                    s = re.sub(r'[^a-z0-9]+', '_', s).strip('_')
                    return s

                # Détection auto de la ligne d'en-tête (cherche "n° facture" ou similaire)
                # On lit SANS dtype=str pour préserver les types datetime des dates Excel,
                # puis on convertit en str colonne par colonne après détection.
                df_map = None
                detected_header = None
                for header_row in range(0, 5):
                    try:
                        if map_file.name.endswith('.csv'):
                            tmp = pd.read_csv(map_file, dtype=str, header=header_row)
                        else:
                            map_file.seek(0)
                            tmp = pd.read_excel(map_file, header=header_row)
                        norm_cols = [normalize(c) for c in tmp.columns]
                        if any(c in aliases_piece for c in norm_cols) and \
                           any(c in aliases_ref for c in norm_cols):
                            df_map = tmp
                            df_map.columns = norm_cols
                            detected_header = header_row
                            break
                    except Exception:
                        continue

                if df_map is None:
                    st.error("Impossible de détecter les colonnes `facture` et `dossier` "
                             "dans le fichier. Vérifiez les en-têtes.")
                else:
                    col_piece = next(c for c in aliases_piece if c in df_map.columns)
                    col_ref = next(c for c in aliases_ref if c in df_map.columns)
                    col_comp = next((c for c in aliases_comp if c in df_map.columns), None)
                    col_date = next((c for c in aliases_date if c in df_map.columns), None)
                    col_situation = next(
                        (c for c in aliases_situation if c in df_map.columns), None)

                    st.info(f"✅ En-tête détecté en ligne {detected_header + 1}. "
                            f"Colonnes utilisées : facture=`{col_piece}`, "
                            f"dossier=`{col_ref}`"
                            + (f", client=`{col_comp}`" if col_comp else "")
                            + (f", date=`{col_date}`" if col_date else "")
                            + (f", situation=`{col_situation}`" if col_situation else ""))

                    def _fmt_date(v):
                        """Parse une date (Timestamp, serial Excel, str DD/MM/YYYY) vers YYYY-MM-DD."""
                        if v is None or pd.isna(v):
                            return ''
                        # Cas 1 : déjà un Timestamp/datetime (lecture Excel sans dtype=str)
                        if isinstance(v, (pd.Timestamp, datetime)):
                            return v.date().isoformat() if hasattr(v, 'date') \
                                else pd.Timestamp(v).date().isoformat()
                        s = str(v).strip()
                        if s == '' or s.lower() == 'nan':
                            return ''
                        # Cas 2 : numéro de série Excel (ex 45870 = 01/08/2025)
                        try:
                            n = float(s.replace(',', '.'))
                            if 20000 < n < 80000:
                                d = pd.Timestamp('1899-12-30') + pd.Timedelta(days=n)
                                return d.date().isoformat()
                        except (ValueError, TypeError):
                            pass
                        # Cas 3 : chaîne — on tente avec dayfirst=True
                        try:
                            d = pd.to_datetime(s, errors='coerce', dayfirst=True)
                            if pd.isna(d):
                                return ''
                            return d.date().isoformat()
                        except Exception:
                            return ''

                    # Construit un index des refs CRM : gère les dossiers regroupés (ex CRM "830/831")
                    # et normalise les zéros de tête (ex PROGEMI "830" <-> CRM "00830").
                    # Chaque sous-ref d'un dossier CRM groupé pointe vers la ref CRM complète.
                    df_d_current = read_sheet('dossiers')
                    crm_refs = df_d_current['ref_client'].astype(str).tolist() \
                        if not df_d_current.empty and 'ref_client' in df_d_current.columns else []
                    import re as _re
                    crm_index = {}  # clé = sous-ref normalisée sans zéros, valeur = ref CRM d'origine
                    for cr in crm_refs:
                        cr_clean = cr.strip()
                        if not cr_clean:
                            continue
                        # Éclate les dossiers groupés côté CRM et indexe chaque sous-ref
                        for sub in _re.split(r'[/,;]+', cr_clean):
                            sub = sub.strip()
                            if sub:
                                crm_index[sub.lstrip('0') or '0'] = cr_clean
                        # Indexe aussi la ref complète telle quelle (au cas où PROGEMI contient "830/831")
                        crm_index[cr_clean.lstrip('0') or '0'] = cr_clean

                    def resolve_ref(raw):
                        """Renvoie la ref CRM (éventuellement groupée) si trouvée, sinon la ref brute."""
                        raw = str(raw).strip()
                        if not raw:
                            return ''
                        key = raw.lstrip('0') or '0'
                        return crm_index.get(key, raw)

                    rows = []
                    nb_groupes_crm = sum(1 for cr in crm_refs if _re.search(r'[/,;]', cr or ''))
                    nb_non_resolus = 0
                    crm_values = set(crm_index.values())
                    for _, r in df_map.iterrows():
                        pr = to_str(r.get(col_piece, ''))
                        rc_raw = to_str(r.get(col_ref, ''))
                        if not pr or not rc_raw or pr.lower() == 'nan':
                            continue
                        resolved = resolve_ref(rc_raw)
                        if resolved not in crm_values:
                            nb_non_resolus += 1
                        rows.append({
                            'piece_ref': pr,
                            'ref_client': resolved,
                            'comp_aux_num': to_str(r.get(col_comp, '')) if col_comp else '',
                            'date_facture': _fmt_date(r.get(col_date, '')) if col_date else '',
                            'situation': to_str(r.get(col_situation, '')) if col_situation else '',
                        })

                    new_df = pd.DataFrame(rows)
                    if nb_groupes_crm:
                        st.info(f"🔀 {nb_groupes_crm} dossiers CRM regroupés détectés "
                                f"(ex: `830/831`) — les factures PROGEMI `830` et `831` "
                                f"pointeront toutes vers le dossier groupé")
                    if nb_non_resolus and crm_values:
                        st.warning(f"⚠️ {nb_non_resolus} refs PROGEMI non trouvées dans le CRM "
                                   f"(importez d'abord le CRM ou vérifiez les codes)")

                    if mode.startswith("Ajouter"):
                        existing = read_sheet('mapping')
                        if not existing.empty:
                            # Upsert sur clé normalisée (gère 23/265 vs 23/0000265)
                            def _np(s):
                                s = str(s or '').strip()
                                if not s:
                                    return ''
                                return '/'.join(p.lstrip('0') or '0' for p in s.split('/'))
                            new_keys = set(new_df['piece_ref'].apply(_np))
                            existing = existing[~existing['piece_ref']
                                                 .apply(_np).isin(new_keys)]
                            merged = pd.concat([existing, new_df], ignore_index=True)
                        else:
                            merged = new_df
                        replace_sheet('mapping', merged)
                        st.success(f"✅ {len(rows)} correspondances importées "
                                   f"(total : {len(merged)})")
                    else:
                        replace_sheet('mapping', new_df)
                        st.success(f"✅ {len(rows)} correspondances facture → dossier importées")
            except Exception as e:
                st.error(f"Erreur : {e}")

        st.divider()
        st.markdown("**Affectation manuelle** — factures non rattachées à un dossier")
        st.caption("Les choix sont sauvegardés dans Google Sheets et persistent à chaque import. "
                   "Marquez une facture « Hors CRM » pour qu'elle ne réapparaisse plus.")

        df_c = read_sheet('creances')
        df_m = read_sheet('mapping')
        df_d = read_sheet('dossiers')

        # Même fonction de normalisation que dans load_creances_enrichies
        def _norm_piece(s):
            s = str(s).strip()
            if not s:
                return ''
            return '/'.join(p.lstrip('0') or '0' for p in s.split('/'))

        if not df_c.empty and not df_d.empty:
            df_c['debit'] = pd.to_numeric(df_c['debit'], errors='coerce').fillna(0)
            df_c['credit'] = pd.to_numeric(df_c['credit'], errors='coerce').fillna(0)
            df_c['solde'] = df_c['debit'] - df_c['credit']
            df_c = df_c[(df_c['ecriture_let'].isna()) | (df_c['ecriture_let'] == '')]
            df_c = df_c[df_c['solde'] > 0]

            # Compare sur clé normalisée pour gérer 22/1 vs 22/0000001
            mapped_keys = set(df_m['piece_ref'].apply(_norm_piece).tolist()) \
                if not df_m.empty else set()
            df_c['_pk'] = df_c['piece_ref'].apply(_norm_piece)
            non_map = df_c[~df_c['_pk'].isin(mapped_keys)]
            # Dédup sur (piece_ref, comp_aux_num) pour éviter les doublons 411 vs 416
            non_map = non_map.groupby(['piece_ref', 'comp_aux_num'], dropna=False).agg(
                comp_aux_lib=('comp_aux_lib', 'first'),
                solde=('solde', 'sum'),
                date=('ecriture_date', 'first')
            ).reset_index().sort_values('solde', ascending=False)

            if non_map.empty:
                st.success("✅ Toutes les factures ouvertes sont rattachées (ou marquées Hors CRM).")
            else:
                total_non_map = non_map['solde'].sum()
                c_a, c_b = st.columns(2)
                c_a.metric("Factures à traiter", len(non_map))
                c_b.metric("Montant concerné", f"{total_non_map:,.0f} €".replace(",", " "))

                # Auto-classification Hors CRM par motif (préfixe de piece_ref)
                with st.expander("⚡ Auto-classer des factures comme Hors CRM (par motif)"):
                    import re as _re2
                    pattern = st.text_input(
                        "Motif regex sur le n° de facture",
                        value=r"^FC",
                        help="Exemples : `^FC` pour toutes les factures commençant par FC, "
                             "`^(FC|AV)` pour FC ou AV, `^FC\\d+$` pour FC suivi de chiffres uniquement"
                    )
                    try:
                        rx = _re2.compile(pattern, _re2.IGNORECASE)
                        preview = non_map[non_map['piece_ref'].apply(
                            lambda x: bool(rx.search(str(x))))]
                    except _re2.error as e:
                        st.error(f"Motif invalide : {e}")
                        preview = pd.DataFrame()

                    if not preview.empty:
                        st.write(f"**{len(preview)} factures** correspondraient au motif "
                                 f"(total : {preview['solde'].sum():,.0f} €)".replace(",", " "))
                        st.dataframe(
                            preview[['piece_ref', 'comp_aux_lib', 'solde']].head(10),
                            use_container_width=True, hide_index=True
                        )
                        if st.button(f"⊘ Marquer ces {len(preview)} factures Hors CRM",
                                     type="primary", key="btn_auto_hors"):
                            new_rows = pd.DataFrame([{
                                'piece_ref': r['piece_ref'],
                                'ref_client': '__HORS_CRM__',
                                'comp_aux_num': r['comp_aux_num'],
                            } for _, r in preview.iterrows()])
                            # retire d'abord les éventuels mappings existants sur ces pieces
                            keys_to_remove = set(preview['piece_ref'].apply(_norm_piece))
                            existing = df_m[~df_m['piece_ref'].apply(_norm_piece)
                                             .isin(keys_to_remove)] \
                                if not df_m.empty else pd.DataFrame(columns=HEADERS['mapping'])
                            merged = pd.concat([existing, new_rows], ignore_index=True)
                            replace_sheet('mapping', merged)
                            st.success(f"✅ {len(preview)} factures marquées Hors CRM")
                            st.rerun()
                    elif pattern:
                        st.info("Aucune facture ne correspond à ce motif.")

                # Recherche + pagination
                q = st.text_input("🔎 Rechercher (n° facture, client, code compta)",
                                  key="search_nonmap").strip().lower()
                if q:
                    mask = (non_map['piece_ref'].str.lower().str.contains(q, na=False)
                            | non_map['comp_aux_lib'].str.lower().str.contains(q, na=False)
                            | non_map['comp_aux_num'].str.lower().str.contains(q, na=False))
                    non_map = non_map[mask]

                per_page = 25
                nb_pages = max(1, (len(non_map) + per_page - 1) // per_page)
                page = st.number_input(f"Page (1 à {nb_pages})", min_value=1,
                                       max_value=nb_pages, value=1, step=1,
                                       key="page_nonmap")
                start = (page - 1) * per_page
                page_df = non_map.iloc[start:start + per_page]

                # Options dossiers CRM + option spéciale Hors CRM
                HORS_CRM = "__HORS_CRM__"
                options_labels = ["— Choisir un dossier —",
                                  "⊘ Hors CRM (facture sans dossier)"]
                options_vals = ["", HORS_CRM]
                for _, r in df_d.iterrows():
                    if r['ref_client']:
                        options_labels.append(f"{r['ref_client']} — {r['client']}")
                        options_vals.append(r['ref_client'])

                st.divider()
                for i_row, row in page_df.reset_index(drop=True).iterrows():
                    c1, c2, c3 = st.columns([2, 3, 1])
                    c1.write(f"**{row['piece_ref']}**")
                    c1.caption(f"{row['comp_aux_num']} — {row['comp_aux_lib']}")
                    # Clé unique : piece_ref + comp_aux_num + index (évite les doublons 411/416)
                    sel_key = f"map_{row['piece_ref']}_{row['comp_aux_num']}_{i_row}"
                    idx = c2.selectbox(
                        "Dossier", range(len(options_labels)),
                        format_func=lambda i: options_labels[i],
                        key=sel_key,
                        label_visibility="collapsed")
                    c3.write(f"{row['solde']:,.0f} €".replace(",", " "))
                    if idx > 0:  # 0 = placeholder "— Choisir —"
                        ref_val = options_vals[idx]
                        new_row = pd.DataFrame([{
                            'piece_ref': row['piece_ref'],
                            'ref_client': ref_val,
                            'comp_aux_num': row['comp_aux_num'],
                        }])
                        existing = df_m[df_m['piece_ref'].apply(_norm_piece)
                                        != _norm_piece(row['piece_ref'])] \
                            if not df_m.empty else pd.DataFrame(columns=HEADERS['mapping'])
                        df_m_updated = pd.concat([existing, new_row], ignore_index=True)
                        replace_sheet('mapping', df_m_updated)
                        st.rerun()

            # --- Section Hors CRM : date de facture manuelle + réaffectation ---
            if not df_m.empty and (df_m['ref_client'] == '__HORS_CRM__').any():
                with st.expander("⊘ Factures marquées Hors CRM "
                                 "(saisir la date de facture / réaffecter)"):
                    hors_all = df_m[df_m['ref_client'] == '__HORS_CRM__']

                    # Référentiel des factures encore dues (net FIFO)
                    enr_open = load_creances_enrichies(only_open=True)
                    enr_open = enr_open[enr_open['solde'].abs() > 0.01]
                    open_pk = set(enr_open['piece_ref'].apply(_norm_piece)) \
                        if not enr_open.empty else set()

                    # Ne garde que les Hors CRM encore dues, dédupliquées par n°
                    hors = hors_all[hors_all['piece_ref'].apply(_norm_piece)
                                    .isin(open_pk)].copy()
                    # Dédoublonnage par n° de facture : garde la ligne qui a une
                    # date_facture si elle existe (évite d'afficher un doublon vide)
                    if 'date_facture' in hors.columns:
                        hors['_hasdate'] = hors['date_facture'].fillna('') \
                            .astype(str).str.strip().ne('').astype(int)
                        hors = hors.sort_values('_hasdate', ascending=False)
                    hors['_pkn'] = hors['piece_ref'].apply(_norm_piece)
                    hors = hors.drop_duplicates('_pkn')
                    nb_soldees = len(hors_all) - len(
                        hors_all[hors_all['piece_ref'].apply(_norm_piece)
                                 .isin(open_pk)])

                    st.write(f"{len(hors)} facture(s) Hors CRM encore due(s)")
                    if nb_soldees:
                        cinfo, cbtn = st.columns([3, 1])
                        cinfo.caption(f"🧾 {nb_soldees} facture(s) Hors CRM soldée(s) "
                                      f"masquée(s).")
                        if cbtn.button("🧹 Purger les soldées",
                                       key="purge_hors_soldees",
                                       help="Retire définitivement les Hors CRM "
                                            "soldées de la feuille mapping"):
                            keep = df_m[~(
                                (df_m['ref_client'] == '__HORS_CRM__')
                                & (~df_m['piece_ref'].apply(_norm_piece)
                                   .isin(open_pk)))]
                            replace_sheet('mapping', keep)
                            st.success(f"✅ {nb_soldees} facture(s) soldée(s) purgée(s)")
                            st.rerun()
                    st.caption("💡 Saisissez la date de facture pour les anciennes "
                               "factures (utile après clôture comptable, quand la "
                               "date de pièce du FEC repasse au 01/01).")
                    for i_h, hr in hors.iterrows():
                        cc1, cc2, cc3, cc4 = st.columns([3, 2, 1, 1])
                        cc1.write(f"**{hr['piece_ref']}**")
                        cc1.caption(f"{hr['comp_aux_num']}")
                        # Date actuelle si déjà saisie
                        cur_df = None
                        _dfv = str(hr.get('date_facture', '') or '')
                        if _dfv and _dfv.lower() != 'nan':
                            try:
                                cur_df = pd.to_datetime(
                                    _dfv, errors='coerce', dayfirst=True).date()
                            except Exception:
                                cur_df = None
                        new_df_date = cc2.date_input(
                            "Date facture", value=cur_df,
                            key=f"horsdate_{i_h}_{hr['piece_ref']}",
                            format="DD/MM/YYYY", label_visibility="collapsed")
                        if cc3.button("💾", key=f"savehd_{i_h}_{hr['piece_ref']}",
                                      help="Enregistrer la date"):
                            df_m_upd = df_m.copy()
                            if 'date_facture' not in df_m_upd.columns:
                                df_m_upd['date_facture'] = ''
                            # Applique la date à TOUTES les lignes du même n°
                            # (gère les éventuels doublons de mapping)
                            pk_target = _norm_piece(hr['piece_ref'])
                            mask_pk = df_m_upd['piece_ref'].apply(_norm_piece) == pk_target
                            df_m_upd.loc[mask_pk, 'date_facture'] = \
                                new_df_date.isoformat() if new_df_date else ''
                            replace_sheet('mapping', df_m_upd)
                            st.success(f"Date enregistrée pour {hr['piece_ref']}")
                            st.rerun()
                        if cc4.button("↶", key=f"unhors_{i_h}_{hr['piece_ref']}",
                                      help="Annuler le statut Hors CRM"):
                            pk_t = _norm_piece(hr['piece_ref'])
                            df_m_cleaned = df_m[df_m['piece_ref'].apply(_norm_piece)
                                                != pk_t]
                            replace_sheet('mapping', df_m_cleaned)
                            st.rerun()

    with tab4:
        st.markdown("**Gestion des dossiers en contentieux**")
        st.caption("Les dossiers listés ici sont exclus de l'export commerciaux "
                   "et apparaissent dans un export dédié.")

        df_ct = read_sheet('contentieux')
        df_d_ct = read_sheet('dossiers')
        df_c_ct = read_sheet('creances')

        # --- Formulaire d'ajout ---
        st.markdown("### Ajouter un dossier au contentieux")

        # Construit la liste des dossiers disponibles (CRM + clients FEC sans dossier)
        already = set(df_ct['ref_client'].tolist()) if not df_ct.empty else set()

        options_add_labels = ["— Choisir un dossier —"]
        options_add_vals = [None]

        if not df_d_ct.empty:
            for _, r in df_d_ct.iterrows():
                if r['ref_client'] and r['ref_client'] not in already:
                    options_add_labels.append(
                        f"CRM — {r['ref_client']} — {r['client']}")
                    options_add_vals.append({
                        'ref_client': r['ref_client'],
                        'comp_aux_num': '',
                    })

        # Permet aussi d'ajouter un client FEC sans dossier CRM (via comp_aux_num)
        if not df_c_ct.empty:
            clients_fec = df_c_ct[['comp_aux_num', 'comp_aux_lib']] \
                .drop_duplicates('comp_aux_num')
            for _, r in clients_fec.iterrows():
                label = f"FEC — {r['comp_aux_num']} — {r['comp_aux_lib']}"
                key_val = f"FEC:{r['comp_aux_num']}"
                if key_val not in already:
                    options_add_labels.append(label)
                    options_add_vals.append({
                        'ref_client': key_val,
                        'comp_aux_num': r['comp_aux_num'],
                    })

        c_a, c_b, c_c = st.columns([3, 2, 3])
        idx_sel = c_a.selectbox("Dossier", range(len(options_add_labels)),
                                format_func=lambda i: options_add_labels[i],
                                key="ct_add_dossier")
        resp = c_b.text_input("Responsable", key="ct_add_resp",
                              placeholder="Nom du gestionnaire")
        comm = c_c.text_input("Commentaire", key="ct_add_comm",
                              placeholder="Facultatif")

        c_d, c_e = st.columns(2)
        prov_r = c_d.number_input("Provision pour risque (€)",
                                  min_value=0.0, step=100.0, value=0.0,
                                  key="ct_add_prov_r")
        prov_cd = c_e.number_input("Provision créances douteuses (€)",
                                   min_value=0.0, step=100.0, value=0.0,
                                   key="ct_add_prov_cd")

        if st.button("➕ Ajouter au contentieux", type="primary"):
            if idx_sel == 0:
                st.warning("Sélectionnez un dossier.")
            elif not resp.strip():
                st.warning("Le responsable est obligatoire.")
            else:
                payload = options_add_vals[idx_sel]
                new_row = pd.DataFrame([{
                    'ref_client': payload['ref_client'],
                    'comp_aux_num': payload['comp_aux_num'],
                    'responsable': resp.strip(),
                    'date_passage': datetime.now().date().isoformat(),
                    'commentaire': comm.strip(),
                    'provision_risque': prov_r,
                    'provision_creances_douteuses': prov_cd,
                }])
                merged = pd.concat([df_ct, new_row], ignore_index=True) \
                    if not df_ct.empty else new_row
                replace_sheet('contentieux', merged)
                st.success("✅ Ajouté au contentieux.")
                st.rerun()

        # --- Liste des dossiers en contentieux (édition provisions) ---
        st.markdown("### Dossiers actuellement en contentieux")
        if df_ct.empty:
            st.info("Aucun dossier en contentieux.")
        else:
            st.caption(f"{len(df_ct)} dossier(s) — éditez les provisions ci-dessous puis "
                       "cliquez **Enregistrer**.")

            # Forme un dataframe éditable
            df_ct_edit = df_ct.copy()
            for col in ['provision_risque', 'provision_creances_douteuses']:
                if col not in df_ct_edit.columns:
                    df_ct_edit[col] = 0
                df_ct_edit[col] = pd.to_numeric(df_ct_edit[col], errors='coerce').fillna(0)

            edited = st.data_editor(
                df_ct_edit[['ref_client', 'responsable', 'date_passage',
                            'commentaire', 'provision_risque',
                            'provision_creances_douteuses']],
                column_config={
                    'ref_client': st.column_config.TextColumn('Dossier', disabled=True),
                    'responsable': st.column_config.TextColumn('Responsable'),
                    'date_passage': st.column_config.TextColumn('Date passage', disabled=True),
                    'commentaire': st.column_config.TextColumn('Commentaire'),
                    'provision_risque': st.column_config.NumberColumn(
                        'Prov. risque (€)', min_value=0, step=100, format="%.2f"),
                    'provision_creances_douteuses': st.column_config.NumberColumn(
                        'Prov. créances douteuses (€)',
                        min_value=0, step=100, format="%.2f"),
                },
                use_container_width=True, hide_index=True, key="ct_editor",
                num_rows="fixed",
            )

            colb1, colb2 = st.columns([1, 5])
            if colb1.button("💾 Enregistrer", type="primary"):
                # Réinjecte les valeurs éditées dans df_ct (en gardant comp_aux_num)
                df_ct_new = df_ct.copy()
                for c in ['responsable', 'commentaire', 'provision_risque',
                          'provision_creances_douteuses']:
                    df_ct_new[c] = edited[c].values
                replace_sheet('contentieux', df_ct_new)
                st.success("✅ Modifications enregistrées.")
                st.rerun()

            st.markdown("**Retirer un dossier du contentieux :**")
            for i, r in df_ct.iterrows():
                cc1, cc2 = st.columns([5, 1])
                cc1.write(f"{r['ref_client']} — 👤 {r['responsable']}")
                if cc2.button("🗑️", key=f"del_ct_{i}", help="Retirer du contentieux"):
                    df_ct_cleaned = df_ct.drop(index=i).reset_index(drop=True)
                    replace_sheet('contentieux', df_ct_cleaned)
                    st.rerun()

    with tab5:
        st.markdown("**Consignations chez l'huissier (chantiers livrés)**")
        st.caption("Saisie manuelle des montants consignés. Soustraits du solde "
                   "dans la synthèse 'Chantiers livrés'.")

        df_cons_e = read_sheet('consignations')
        df_c_cons = read_sheet('creances')

        st.markdown("### Ajouter / mettre à jour une consignation")

        # Liste des clients FEC pour saisie
        if df_c_cons.empty:
            st.info("Importez d'abord le FEC.")
        else:
            clients_list = df_c_cons[['comp_aux_num', 'comp_aux_lib']] \
                .drop_duplicates('comp_aux_num') \
                .sort_values('comp_aux_lib')
            options_lab = ["— Choisir un client —"] + [
                f"{r['comp_aux_num']} — {r['comp_aux_lib']}"
                for _, r in clients_list.iterrows()
            ]
            options_val = [None] + clients_list['comp_aux_num'].tolist()

            cc1, cc2, cc3 = st.columns([3, 2, 3])
            sel_idx = cc1.selectbox("Client", range(len(options_lab)),
                                     format_func=lambda i: options_lab[i],
                                     key="cons_sel")
            mt = cc2.number_input("Montant consigné (€)", min_value=0.0,
                                   step=100.0, format="%.2f", key="cons_mt")
            comm = cc3.text_input("Commentaire", key="cons_comm",
                                   placeholder="Facultatif")

            if st.button("💾 Enregistrer la consignation", type="primary"):
                if sel_idx == 0:
                    st.warning("Sélectionnez un client.")
                else:
                    comp = options_val[sel_idx]
                    new_row = pd.DataFrame([{
                        'comp_aux_num': comp,
                        'ref_client': '',
                        'montant_consigne': float(mt),
                        'date_consignation': datetime.now().date().isoformat(),
                        'commentaire': comm.strip(),
                    }])
                    if df_cons_e.empty:
                        merged = new_row
                    else:
                        merged = df_cons_e[df_cons_e['comp_aux_num'] != comp]
                        merged = pd.concat([merged, new_row], ignore_index=True)
                    replace_sheet('consignations', merged)
                    st.success("✅ Consignation enregistrée.")
                    st.rerun()

        st.markdown("### Consignations enregistrées")
        if df_cons_e.empty:
            st.info("Aucune consignation.")
        else:
            for i_c, r in df_cons_e.iterrows():
                cc1, cc2, cc3 = st.columns([3, 2, 1])
                cc1.write(f"**{r['comp_aux_num']}** — {r.get('commentaire', '') or '—'}")
                cc1.caption(f"Le {r.get('date_consignation', '')}")
                cc2.write(f"💰 {float(r.get('montant_consigne', 0) or 0):,.2f} €"
                          .replace(",", " "))
                if cc3.button("🗑️", key=f"del_cons_{i_c}_{r['comp_aux_num']}",
                              help="Supprimer"):
                    df_cons_cleaned = df_cons_e.drop(index=i_c).reset_index(drop=True)
                    replace_sheet('consignations', df_cons_cleaned)
                    st.rerun()

    with tab6:
        st.markdown("**Gestion des utilisateurs**")
        st.caption("Les utilisateurs ajoutés ici peuvent être sélectionnés "
                   "dans la sidebar et recevoir des tâches assignées.")

        df_u = read_sheet('users')

        st.markdown("### Ajouter / mettre à jour un utilisateur")
        cu1, cu2, cu3 = st.columns([3, 3, 1])
        new_email = cu1.text_input("Email", key="u_email",
                                    placeholder="nom@designconstructions.com")
        new_nom = cu2.text_input("Nom affiché", key="u_nom",
                                  placeholder="Prénom Nom")
        new_actif = cu3.selectbox("Actif", ["oui", "non"], key="u_actif")

        if st.button("➕ Enregistrer", type="primary"):
            if not new_email.strip() or '@' not in new_email:
                st.warning("Email invalide.")
            elif not new_nom.strip():
                st.warning("Le nom est obligatoire.")
            else:
                new_row = pd.DataFrame([{
                    'email': new_email.strip().lower(),
                    'nom_affichage': new_nom.strip(),
                    'actif': new_actif,
                }])
                if df_u.empty:
                    merged = new_row
                else:
                    # Upsert sur email
                    merged = df_u[df_u['email'].astype(str).str.lower()
                                  != new_email.strip().lower()]
                    merged = pd.concat([merged, new_row], ignore_index=True)
                replace_sheet('users', merged)
                st.success(f"✅ Utilisateur enregistré : {new_nom}")
                st.rerun()

        st.markdown("### Utilisateurs enregistrés")
        if df_u.empty:
            st.info("Aucun utilisateur. Ajoutez-en un avec le formulaire ci-dessus.")
        else:
            for i_u, r in df_u.iterrows():
                cc1, cc2, cc3, cc4, cc5 = st.columns([3, 2, 1, 1, 1])
                cc1.write(f"**{r.get('nom_affichage', '')}**")
                cc1.caption(r.get('email', ''))
                actif = str(r.get('actif', 'oui')).lower() in (
                    'oui', 'true', '1', 'yes', '')
                has_pwd = bool(str(r.get('password_hash', '') or '').strip())
                cc2.write(("🟢 Actif" if actif else "🔴 Inactif")
                          + (" · 🔑" if has_pwd else " · ⚠️ sans mdp"))
                if cc3.button("🔄", key=f"toggle_u_{i_u}",
                              help="Basculer actif/inactif"):
                    df_u.loc[i_u, 'actif'] = 'non' if actif else 'oui'
                    replace_sheet('users', df_u)
                    st.rerun()
                if cc4.button("🔐", key=f"reset_u_{i_u}",
                              help="Réinitialiser le mot de passe"):
                    st.session_state[f'reset_pwd_for_{i_u}'] = True
                if cc5.button("🗑️", key=f"del_u_{i_u}", help="Supprimer"):
                    df_u_cleaned = df_u.drop(index=i_u).reset_index(drop=True)
                    replace_sheet('users', df_u_cleaned)
                    st.rerun()

                # Formulaire de réinitialisation inline
                if st.session_state.get(f'reset_pwd_for_{i_u}'):
                    with st.form(f"reset_form_{i_u}"):
                        new_pwd = st.text_input(
                            f"Nouveau mot de passe pour {r.get('email', '')}",
                            type="password", key=f"reset_pwd_input_{i_u}")
                        cra, crb = st.columns(2)
                        if cra.form_submit_button("💾 Définir", type="primary"):
                            if len(new_pwd) < 6:
                                st.warning("Le mot de passe doit faire "
                                           "au moins 6 caractères.")
                            else:
                                df_u.loc[i_u, 'password_hash'] = hash_pwd(new_pwd)
                                replace_sheet('users', df_u)
                                st.session_state[f'reset_pwd_for_{i_u}'] = False
                                st.success("Mot de passe défini.")
                                st.rerun()
                        if crb.form_submit_button("Annuler"):
                            st.session_state[f'reset_pwd_for_{i_u}'] = False
                            st.rerun()

    st.divider()
    df_c = read_sheet('creances')
    df_d = read_sheet('dossiers')
    df_m = read_sheet('mapping')
    n_cr = len(df_c)
    n_cr_ouv = 0
    if not df_c.empty and 'ecriture_let' in df_c.columns:
        n_cr_ouv = len(df_c[(df_c['ecriture_let'].isna()) | (df_c['ecriture_let'] == '')])
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Écritures FEC", n_cr)
    c2.metric("Non lettrées", n_cr_ouv)
    c3.metric("Dossiers CRM", len(df_d))
    c4.metric("Mappings", len(df_m))


def page_creances():
    st.header("📊 Créances ouvertes")

    df = load_creances_enrichies(only_open=True)
    if df.empty:
        st.warning("Aucune créance. Importez le FEC dans l'onglet Import.")
        return

    df = df[df['solde'].abs() > 0.01]

    if 'conducteur' not in df.columns:
        df['conducteur'] = ''

    c1, c2, c3, c4 = st.columns(4)
    coms = ['(Tous)'] + sorted([c for c in df['commercial'].dropna().unique() if c])
    filt_com = c1.selectbox("Commercial", coms)
    conds = ['(Tous)'] + sorted([c for c in df['conducteur'].dropna().unique() if c])
    filt_cond = c2.selectbox("Conducteur", conds)
    etats = ['(Tous)'] + sorted([e for e in df['etat'].dropna().unique() if e])
    filt_et = c3.selectbox("État dossier", etats)
    seuil = c4.number_input("Solde mini (€)", value=0, step=500)

    f = df.copy()
    if filt_com != '(Tous)':
        f = f[f['commercial'] == filt_com]
    if filt_cond != '(Tous)':
        f = f[f['conducteur'] == filt_cond]
    if filt_et != '(Tous)':
        f = f[f['etat'] == filt_et]
    f = f[f['solde'] >= seuil]

    # Sépare les dossiers en contentieux du reste
    f_contentieux = f[f['contentieux']].copy() if 'contentieux' in f.columns else pd.DataFrame()
    f = f[~f['contentieux']] if 'contentieux' in f.columns else f

    total = f['solde'].sum()
    nb_cli = f['comp_aux_num'].nunique()
    nb_mappes = f['ref_client'].fillna('').astype(bool).sum()
    non_mappe = f[f['ref_client'].fillna('') == '']['solde'].sum()

    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Solde total dû", f"{total:,.0f} €".replace(",", " "))
    k2.metric("Clients concernés", nb_cli)
    k3.metric("Lignes rattachées", f"{nb_mappes} / {len(f)}")
    k4.metric("Non rattaché (€)", f"{non_mappe:,.0f}".replace(",", " "))

    # Coloration conditionnelle jours de retard (charte DCA) :
    # vert <7, orange 7-29, rouge >=30
    def _color_retard(v):
        try:
            v = int(v)
        except Exception:
            return ''
        if v <= 0:
            return ''
        if v < 7:
            return 'background-color: #C0DD97; color: #355A10;'  # vert DCA
        if v < 30:
            return 'background-color: #F5D7A8; color: #8B5A00;'  # orange DCA
        return 'background-color: #F5BEB6; color: #7A1F12;'  # rouge DCA

    # Split en 2 : chantiers en cours / chantiers livrés
    f_en_cours = f[~f.get('est_livre', False)] if 'est_livre' in f.columns else f
    f_livres = f[f.get('est_livre', False)] if 'est_livre' in f.columns \
        else f.iloc[0:0]

    # --- Synthèse Chantiers EN COURS ---
    st.subheader("🏗️ Chantiers en cours")
    if f_en_cours.empty:
        st.caption("Aucun chantier en cours.")
    else:
        synth_ec = f_en_cours.groupby(
            ['comp_aux_num', 'comp_aux_lib', 'ref_client',
             'commercial', 'conducteur', 'etat'], dropna=False).agg(
            solde=('solde', 'sum'),
            nb=('piece_ref', 'count'),
            derniere_facture=('date_facture_eff', 'max'),
            jours_retard=('jours_retard', 'max')
        ).reset_index().sort_values('solde', ascending=False)

        ka, kb = st.columns(2)
        ka.metric("Total dû en cours",
                  f"{f_en_cours['solde'].sum():,.0f} €".replace(",", " "))
        kb.metric("Nb dossiers", len(synth_ec))

        synth_ec['derniere_facture'] = fr_series(synth_ec['derniere_facture'])
        synth_ec_display = synth_ec.rename(columns={
            'comp_aux_num': 'Code compta', 'comp_aux_lib': 'Client',
            'ref_client': 'Ref dossier',
            'commercial': 'Commercial', 'conducteur': 'Conducteur',
            'etat': 'État', 'solde': 'Solde (€)',
            'nb': 'Nb lignes', 'derniere_facture': 'Dernière facture',
            'jours_retard': 'Jours retard'
        })
        styled_ec = synth_ec_display.style.map(_color_retard, subset=['Jours retard']) \
            .format({'Solde (€)': '{:.2f}'})
        st.dataframe(styled_ec, use_container_width=True, hide_index=True)

    # --- Synthèse Chantiers LIVRÉS ---
    st.subheader("🏠 Chantiers livrés")
    if f_livres.empty:
        st.caption("Aucun chantier livré avec créance ouverte.")
    else:
        synth_lv = f_livres.groupby(
            ['comp_aux_num', 'comp_aux_lib', 'ref_client',
             'commercial', 'conducteur', 'etat'], dropna=False).agg(
            solde=('solde', 'sum'),
            nb=('piece_ref', 'count'),
            derniere_facture=('date_facture_eff', 'max'),
            date_livraison=('date_reception', 'first'),
            jours_retard=('jours_retard', 'max'),
            montant_consigne=('montant_consigne', 'first'),
        ).reset_index().sort_values('solde', ascending=False)
        synth_lv['solde_net'] = synth_lv['solde'] - synth_lv['montant_consigne'].fillna(0)

        total_lv = synth_lv['solde'].sum()
        total_cons = synth_lv['montant_consigne'].sum()
        total_net = synth_lv['solde_net'].sum()
        ka, kb, kc, kd = st.columns(4)
        ka.metric("Total dû livrés", f"{total_lv:,.0f} €".replace(",", " "))
        kb.metric("Nb dossiers", len(synth_lv))
        kc.metric("Total consigné", f"{total_cons:,.0f} €".replace(",", " "))
        kd.metric("Solde net", f"{total_net:,.0f} €".replace(",", " "))

        synth_lv['derniere_facture'] = fr_series(synth_lv['derniere_facture'])
        synth_lv['date_livraison'] = fr_series(synth_lv['date_livraison'])
        synth_lv_display = synth_lv.rename(columns={
            'comp_aux_num': 'Code compta', 'comp_aux_lib': 'Client',
            'ref_client': 'Ref dossier',
            'commercial': 'Commercial', 'conducteur': 'Conducteur',
            'etat': 'État', 'solde': 'Solde (€)',
            'montant_consigne': 'Consigné huissier (€)',
            'solde_net': 'Solde net (€)',
            'nb': 'Nb lignes', 'derniere_facture': 'Dernière facture',
            'date_livraison': 'Date livraison',
            'jours_retard': 'Jours retard'
        })
        styled_lv = synth_lv_display.style.map(_color_retard, subset=['Jours retard']) \
            .format({'Solde (€)': '{:.2f}',
                     'Consigné huissier (€)': '{:.2f}',
                     'Solde net (€)': '{:.2f}'})
        st.dataframe(styled_lv, use_container_width=True, hide_index=True)

    # --- Sous-tableau Contentieux ---
    if not f_contentieux.empty:
        st.subheader("⚖️ Dossiers en contentieux")
        synth_ct = f_contentieux.groupby(
            ['comp_aux_num', 'comp_aux_lib', 'ref_client',
             'responsable', 'commercial'], dropna=False).agg(
            solde=('solde', 'sum'),
            nb=('piece_ref', 'count'),
            derniere_facture=('date_facture_eff', 'max'),
            date_livraison=('date_reception', 'first'),
            jours_retard=('jours_retard', 'max'),
            provision_risque=('provision_risque', 'first'),
            provision_creances_douteuses=('provision_creances_douteuses', 'first'),
            montant_consigne=('montant_consigne', 'first'),
        ).reset_index().sort_values('solde', ascending=False)

        total_ct = synth_ct['solde'].sum()
        total_cons_ct = synth_ct['montant_consigne'].sum()
        total_pr = synth_ct['provision_risque'].sum()
        total_pcd = synth_ct['provision_creances_douteuses'].sum()
        ka, kb, kc, kd, ke = st.columns(5)
        ka.metric("Total contentieux", f"{total_ct:,.0f} €".replace(",", " "))
        kb.metric("Dossiers", len(synth_ct))
        kc.metric("Total consigné", f"{total_cons_ct:,.0f} €".replace(",", " "))
        kd.metric("Prov. risque", f"{total_pr:,.0f} €".replace(",", " "))
        ke.metric("Prov. créances douteuses",
                  f"{total_pcd:,.0f} €".replace(",", " "))

        synth_ct['derniere_facture'] = fr_series(synth_ct['derniere_facture'])
        synth_ct['date_livraison'] = fr_series(synth_ct['date_livraison'])
        synth_ct_display = synth_ct.rename(columns={
            'comp_aux_num': 'Code compta', 'comp_aux_lib': 'Client',
            'ref_client': 'Ref dossier',
            'responsable': 'Responsable', 'commercial': 'Commercial',
            'solde': 'Solde (€)', 'nb': 'Nb lignes',
            'derniere_facture': 'Dernière facture',
            'date_livraison': 'Date livraison',
            'jours_retard': 'Jours retard',
            'montant_consigne': 'Consigné huissier (€)',
            'provision_risque': 'Prov. risque (€)',
            'provision_creances_douteuses': 'Prov. créances douteuses (€)',
        })
        styled_ct = synth_ct_display.style.map(_color_retard, subset=['Jours retard']) \
            .format({'Solde (€)': '{:.2f}',
                     'Consigné huissier (€)': '{:.2f}',
                     'Prov. risque (€)': '{:.2f}',
                     'Prov. créances douteuses (€)': '{:.2f}'})
        st.dataframe(styled_ct, use_container_width=True, hide_index=True)

    with st.expander("Détail ligne par ligne"):
        detail = f[['comp_aux_lib', 'piece_ref', 'ecriture_date', 'journal_code',
                    'ecriture_lib', 'debit', 'credit', 'solde', 'commercial', 'etat']].copy()
        detail['ecriture_date'] = fr_series(detail['ecriture_date'])
        st.dataframe(
            detail.rename(columns={
                'comp_aux_lib': 'Client', 'piece_ref': 'Réf. pièce',
                'ecriture_date': 'Date', 'journal_code': 'Journal',
                'ecriture_lib': 'Libellé', 'debit': 'Débit',
                'credit': 'Crédit', 'solde': 'Solde',
                'commercial': 'Commercial', 'etat': 'État'
            }),
            use_container_width=True, hide_index=True
        )


def page_notes():
    st.header("📝 Notes & Relances")

    df_full = load_creances_enrichies(only_open=True)
    if df_full.empty:
        st.warning("Aucune créance ouverte. Importez le FEC.")
        return

    df_full = df_full[df_full['solde'] > 0.01]
    clients = df_full.groupby(['comp_aux_num', 'comp_aux_lib']).agg(
        solde=('solde', 'sum'),
        client=('client', 'first'),
        commercial=('commercial', 'first'),
        ref_client=('ref_client', 'first'),
    ).reset_index().sort_values('solde', ascending=False)

    notes_df = read_sheet('notes')
    if not notes_df.empty:
        notes_df['id'] = pd.to_numeric(notes_df['id'], errors='coerce')

    labels = {f"{r['comp_aux_lib']} — {r['solde']:,.0f} €".replace(",", " "): r['comp_aux_num']
              for _, r in clients.iterrows()}
    sel = st.selectbox("Client", ['— Vue globale —'] + list(labels.keys()))

    if sel == '— Vue globale —':
        st.subheader(f"Toutes les relances ({len(notes_df)})")
        if notes_df.empty:
            st.info("Aucune note.")
        else:
            enriched = notes_df.merge(
                clients[['comp_aux_num', 'client', 'commercial']],
                on='comp_aux_num', how='left')
            st.dataframe(
                enriched[['date_note', 'comp_aux_num', 'client', 'commercial',
                          'auteur', 'action', 'note', 'echeance', 'statut']],
                use_container_width=True, hide_index=True)
        return

    comp_aux_num = labels[sel]
    info = clients[clients['comp_aux_num'] == comp_aux_num].iloc[0]

    c1, c2, c3 = st.columns(3)
    c1.metric("Solde dû", f"{info['solde']:,.2f} €".replace(",", " "))
    c2.write(f"**Client CRM :** {info['client'] or '(non rattaché)'}")
    c3.write(f"**Commercial :** {info['commercial'] or '—'}")

    fac = df_full[df_full['comp_aux_num'] == comp_aux_num][
        ['piece_ref', 'ecriture_date', 'ecriture_lib', 'debit', 'credit', 'solde']]
    with st.expander(f"Détail des {len(fac)} lignes ouvertes"):
        st.dataframe(fac.rename(columns={
            'piece_ref': 'Réf.', 'ecriture_date': 'Date', 'ecriture_lib': 'Libellé',
            'debit': 'Débit', 'credit': 'Crédit', 'solde': 'Solde'
        }), use_container_width=True, hide_index=True)

    st.subheader("Historique des relances")
    client_notes = notes_df[notes_df['comp_aux_num'] == comp_aux_num] \
        if not notes_df.empty else pd.DataFrame()
    if not client_notes.empty:
        client_notes = client_notes.sort_values('date_note', ascending=False)

    if client_notes.empty:
        st.info("Aucune note pour ce client.")
    else:
        active_users_h = get_active_users()
        for _, n in client_notes.iterrows():
            icon = {'Ouvert': '🔴', 'En cours': '🟡', 'Résolu': '🟢'}.get(n['statut'], '⚪')
            nid = int(n['id'])
            with st.expander(f"{icon} {n['date_note']} — {n['auteur']} — {n['action'] or '(note)'}"):
                edit_key = f"edit_mode_{nid}"
                if not st.session_state.get(edit_key, False):
                    # --- Mode lecture ---
                    st.write(n['note'])
                    if n['echeance']:
                        st.caption(f"📅 Échéance : {n['echeance']}")
                    if n.get('assigne_a'):
                        st.caption(f"👤 Assignée à : {n['assigne_a']}")
                    cols = st.columns([2, 1, 1, 1])
                    statuts = ['Ouvert', 'En cours', 'Résolu']
                    cur_idx = statuts.index(n['statut']) if n['statut'] in statuts else 0
                    new_st = cols[0].selectbox("Statut", statuts, index=cur_idx,
                                               key=f"st_{nid}")
                    if cols[1].button("Statut ✓", key=f"up_{nid}"):
                        update_cell_by_id('notes', nid, 'statut', new_st)
                        st.rerun()
                    if cols[2].button("✏️ Modifier", key=f"edbtn_{nid}"):
                        st.session_state[edit_key] = True
                        st.rerun()
                    if cols[3].button("🗑 Supprimer", key=f"del_{nid}"):
                        delete_row_by_id('notes', nid)
                        st.rerun()
                else:
                    # --- Mode édition ---
                    with st.form(f"edit_note_{nid}"):
                        e_action = st.text_input("Type d'action",
                                                  value=n.get('action', '') or '')
                        e_note = st.text_area("Note détaillée",
                                               value=n.get('note', '') or '',
                                               height=120)
                        ec1, ec2, ec3 = st.columns(3)
                        # Échéance
                        cur_ech = None
                        try:
                            if n.get('echeance'):
                                cur_ech = pd.to_datetime(
                                    n['echeance'], dayfirst=True).date()
                        except Exception:
                            cur_ech = None
                        e_ech = ec1.date_input("Échéance", value=cur_ech)
                        statuts = ['Ouvert', 'En cours', 'Résolu']
                        e_idx = statuts.index(n['statut']) \
                            if n['statut'] in statuts else 0
                        e_statut = ec2.selectbox("Statut", statuts, index=e_idx)
                        # Assignation
                        u_opts = ["— Personne —"] + [
                            f"{u['nom']} ({u['email']})" for u in active_users_h]
                        cur_assigne = n.get('assigne_a', '') or ''
                        a_idx = 0
                        for i_u, u in enumerate(active_users_h):
                            if u['email'] == cur_assigne:
                                a_idx = i_u + 1
                                break
                        e_assigne_idx = ec3.selectbox(
                            "Assigner à", range(len(u_opts)),
                            format_func=lambda i: u_opts[i], index=a_idx)

                        fc1, fc2 = st.columns(2)
                        if fc1.form_submit_button("💾 Enregistrer",
                                                   type="primary"):
                            e_assigne = active_users_h[e_assigne_idx - 1]['email'] \
                                if e_assigne_idx > 0 else ''
                            update_row_by_id('notes', nid, {
                                'action': e_action,
                                'note': e_note,
                                'echeance': e_ech.isoformat() if e_ech else '',
                                'statut': e_statut,
                                'assigne_a': e_assigne,
                            })
                            st.session_state[edit_key] = False
                            st.success("Note modifiée.")
                            st.rerun()
                        if fc2.form_submit_button("Annuler"):
                            st.session_state[edit_key] = False
                            st.rerun()

    # --- Résumé direction (un par client, écrasable) ---
    st.subheader("📌 Résumé direction (1 ligne par client)")
    df_res = read_sheet('resumes')
    cur_resume = ''
    cur_action = ''
    cur_resp = ''
    cur_date_recouv = None
    cur_nature = ''
    cur_resume_meta = ''
    if not df_res.empty and 'comp_aux_num' in df_res.columns:
        match = df_res[df_res['comp_aux_num'] == comp_aux_num]
        if not match.empty:
            cur_resume = str(match.iloc[0].get('resume', '') or '')
            cur_action = str(match.iloc[0].get('action_resume', '') or '')
            cur_resp = str(match.iloc[0].get('responsable_action', '') or '')
            cur_nature = str(match.iloc[0].get('nature_creance', '') or '')
            _dr = str(match.iloc[0].get('date_recouvrement', '') or '')
            try:
                if _dr:
                    cur_date_recouv = pd.to_datetime(
                        _dr, errors='coerce', dayfirst=True).date()
            except Exception:
                cur_date_recouv = None
            cur_resume_meta = (f"Mis à jour le {match.iloc[0].get('date_maj', '')} "
                               f"par {match.iloc[0].get('auteur', '')}")

    nat_idx = NATURES_CREANCE.index(cur_nature) \
        if cur_nature in NATURES_CREANCE else 0

    with st.form("resume_form", clear_on_submit=False):
        new_resume = st.text_input(
            "Résumé (max 100 caractères) — repris dans l'export Direction",
            value=cur_resume, max_chars=100,
            placeholder="Ex: Litige sur facture 26/0000123, attente expertise"
        )
        ca, cr = st.columns(2)
        new_action = ca.text_input(
            "Action à mener (max 100 caractères)",
            value=cur_action, max_chars=100,
            placeholder="Ex: Relancer assurance, Saisine huissier..."
        )
        new_resp = cr.text_input(
            "Responsable de l'action",
            value=cur_resp,
            placeholder="Ex: Jean Dupont"
        )
        cn1, cn2 = st.columns(2)
        new_nature = cn1.selectbox(
            "Nature de la créance", NATURES_CREANCE, index=nat_idx)
        new_date_recouv = cn2.date_input(
            "Date de recouvrement estimée", value=cur_date_recouv,
            help="Date à laquelle on estime recevoir le règlement")
        resume_auteur = st.text_input(
            "Auteur (qui rédige ce résumé)",
            value=st.session_state.get('last_auteur', ''),
            key="resume_auteur")
        if cur_resume_meta:
            st.caption(cur_resume_meta)
        cs1, cs2 = st.columns([1, 5])
        if cs1.form_submit_button("💾 Enregistrer le résumé", type="primary"):
            st.session_state['last_auteur'] = resume_auteur
            new_row = pd.DataFrame([{
                'comp_aux_num': comp_aux_num,
                'ref_client': info.get('ref_client', ''),
                'resume': new_resume.strip()[:100],
                'action_resume': new_action.strip()[:100],
                'responsable_action': new_resp.strip(),
                'date_recouvrement': new_date_recouv.isoformat()
                if new_date_recouv else '',
                'nature_creance': new_nature if new_nature != '—' else '',
                'date_maj': datetime.now().strftime('%Y-%m-%d %H:%M'),
                'auteur': resume_auteur,
            }])
            if df_res.empty:
                merged = new_row
            else:
                merged = df_res[df_res['comp_aux_num'] != comp_aux_num]
                merged = pd.concat([merged, new_row], ignore_index=True)
            replace_sheet('resumes', merged)
            st.success("✅ Résumé enregistré.")
            st.rerun()

    st.subheader("➕ Ajouter une note / tâche")
    active_users = get_active_users()
    user_options = ["— Personne —"] + [f"{u['nom']} ({u['email']})"
                                         for u in active_users]
    cur = current_user()
    default_auteur = (cur['nom'] if cur else
                      st.session_state.get('last_auteur', ''))

    with st.form("new_note", clear_on_submit=True):
        c1, c2 = st.columns(2)
        auteur = c1.text_input("Auteur", value=default_auteur)
        action = c2.text_input("Type d'action",
                               placeholder="ex: Appel, Mail, Relance 1...")
        note = st.text_area("Note détaillée", height=100)
        c3, c4, c5 = st.columns(3)
        echeance = c3.date_input("Échéance (optionnel)", value=None)
        statut = c4.selectbox("Statut", ['Ouvert', 'En cours', 'Résolu'])
        assigne_idx = c5.selectbox("Assigner à",
                                     range(len(user_options)),
                                     format_func=lambda i: user_options[i])
        if st.form_submit_button("Enregistrer", type="primary"):
            if note.strip():
                st.session_state['last_auteur'] = auteur
                assigne_email = active_users[assigne_idx - 1]['email'] \
                    if assigne_idx > 0 else ''
                new_id = next_id(notes_df)
                append_row('notes', {
                    'id': new_id,
                    'ref_client': info.get('ref_client', ''),
                    'comp_aux_num': comp_aux_num,
                    'date_note': datetime.now().strftime('%Y-%m-%d %H:%M'),
                    'auteur': auteur,
                    'note': note,
                    'action': action,
                    'echeance': echeance.isoformat() if echeance else '',
                    'statut': statut,
                    'assigne_a': assigne_email,
                })
                st.success("Note enregistrée.")
                st.rerun()


# Charte graphique Design Constructions
DCA_PRIMARY = '60A020'      # vert DCA
DCA_SECONDARY = '2C3E50'    # bleu marine
DCA_NEUTRAL = 'D4880C'      # orange
DCA_BAD = 'C0392B'          # rouge


def _style_header(cell):
    cell.font = Font(name='Segoe UI', bold=True, color='FFFFFF', size=11)
    cell.fill = PatternFill('solid', start_color=DCA_SECONDARY)
    cell.alignment = Alignment(horizontal='center', vertical='center')


def _style_total(cell):
    cell.font = Font(name='Segoe UI', bold=True, color='FFFFFF', size=11)
    cell.fill = PatternFill('solid', start_color=DCA_PRIMARY)
    cell.alignment = Alignment(horizontal='right', vertical='center')


def _autosize(ws):
    for col_cells in ws.columns:
        length = max((len(str(c.value or '')) for c in col_cells), default=10)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = \
            min(max(length + 2, 12), 40)


def page_export():
    st.header("📤 Export")

    df = load_creances_enrichies(only_open=True)
    df = df[df['solde'] > 0.01]
    if df.empty:
        st.warning("Aucune créance à exporter.")
        return

    notes = read_sheet('notes')
    if not notes.empty:
        last_notes = notes.groupby('comp_aux_num').agg(
            derniere_relance=('date_note', 'max'),
            nb_relances=('date_note', 'count')
        ).reset_index()
    else:
        last_notes = pd.DataFrame(columns=['comp_aux_num', 'derniere_relance', 'nb_relances'])

    # Sépare contentieux avant exports
    df_all = df.copy()
    df_ctx = df_all[df_all.get('contentieux', False)] if 'contentieux' in df_all.columns \
        else pd.DataFrame()
    df = df_all[~df_all.get('contentieux', False)] if 'contentieux' in df_all.columns \
        else df_all

    tab1, tab2, tab3, tab4 = st.tabs(["Export commerciaux", "Export Power BI",
                                       "Export Contentieux", "Export Direction"])

    with tab1:
        st.markdown("Classeur Excel : 1 feuille par commercial avec **uniquement les "
                    "chantiers en cours** (livrés et contentieux exclus). "
                    "Détail facture par facture avec jours depuis émission colorés.")

        # Filtre : uniquement chantiers en cours (livrés exclus, contentieux déjà exclu)
        df_ec = df[~df.get('est_livre', False)] if 'est_livre' in df.columns else df

        nb_excl_livres = df['est_livre'].sum() if 'est_livre' in df.columns else 0
        if nb_excl_livres or not df_ctx.empty:
            msgs = []
            if not df_ctx.empty:
                msgs.append(f"{df_ctx['ref_client'].nunique()} dossier(s) contentieux")
            if nb_excl_livres:
                msgs.append(f"{nb_excl_livres} ligne(s) chantier livré")
            st.caption("ℹ️ Exclus de cet export : " + ", ".join(msgs))

        if st.button("🔧 Générer l'export commerciaux", type="primary"):
            wb = openpyxl.Workbook()
            wb.remove(wb.active)

            # Couleurs jours depuis facture (charte DCA) :
            # vert <7, orange 7-15, rouge >15
            def _fill_jours(j):
                try:
                    j = int(j)
                except Exception:
                    return None
                if j <= 0:
                    return None
                if j < 7:
                    return PatternFill('solid', start_color='C0DD97')   # vert
                if j <= 15:
                    return PatternFill('solid', start_color='F5D7A8')   # orange
                return PatternFill('solid', start_color='F5BEB6')        # rouge

            # Une feuille par commercial — uniquement détail des factures non soldées
            for com in sorted(df_ec['commercial'].dropna().unique()):
                if not com:
                    continue
                df_c = df_ec[df_ec['commercial'] == com] \
                    .sort_values(['comp_aux_lib', 'date_facture_eff'])
                if df_c.empty:
                    continue
                safe = com[:31].replace('/', '-').replace('\\', '-')
                ws = wb.create_sheet(safe)
                headers = ['Client', 'Ref dossier', 'N° facture',
                           'Date facture', 'Libellé', 'Solde dû (€)',
                           'Jours depuis facture', 'État']
                ws.append(headers)
                for c in ws[1]:
                    _style_header(c)

                for _, r in df_c.iterrows():
                    jours = int(r.get('jours_retard', 0) or 0)
                    ws.append([
                        r['comp_aux_lib'], r['ref_client'], r['piece_ref'],
                        to_date_obj(r.get('date_facture_eff', '') or r['ecriture_date']),
                        r['ecriture_lib'],
                        round(r['solde'], 2),
                        jours, r['etat']
                    ])
                    fill = _fill_jours(jours)
                    if fill is not None:
                        ws.cell(ws.max_row, 7).fill = fill

                last = ws.max_row + 1
                ws.cell(last, 1, 'TOTAL').font = Font(bold=True)
                ws.cell(last, 6, f'=SUM(F2:F{last - 1})').font = Font(bold=True)
                for row in ws.iter_rows(min_row=2, max_row=last, min_col=6, max_col=6):
                    for c in row:
                        c.number_format = '#,##0.00 €'
                # Format date sur la colonne D (Date facture)
                for row in ws.iter_rows(min_row=2, max_row=last - 1,
                                         min_col=4, max_col=4):
                    for c in row:
                        c.number_format = 'DD/MM/YYYY'
                _autosize(ws)
                ws.freeze_panes = 'A2'

            non_map = df_ec[df_ec['commercial'].fillna('') == '']
            if not non_map.empty:
                ws = wb.create_sheet("Non rattachés")
                ws.append(['Client FEC', 'Code compta', 'Réf. pièce',
                           'Date facture', 'Libellé', 'Solde'])
                for c in ws[1]:
                    _style_header(c)
                for _, r in non_map.iterrows():
                    ws.append([r['comp_aux_lib'], r['comp_aux_num'], r['piece_ref'],
                               to_date_obj(r.get('date_facture_eff', '')
                                           or r['ecriture_date']),
                               r['ecriture_lib'], round(r['solde'], 2)])
                # Format date colonne D
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                         min_col=4, max_col=4):
                    for c in row:
                        c.number_format = 'DD/MM/YYYY'
                _autosize(ws)

            if not notes.empty:
                ws = wb.create_sheet("Relances")
                ws.append(['Date', 'Client', 'Auteur', 'Action', 'Note', 'Échéance', 'Statut'])
                for c in ws[1]:
                    _style_header(c)
                for _, r in notes.iterrows():
                    ws.append([to_date_obj(r['date_note']), r['comp_aux_num'], r['auteur'],
                               r['action'], r['note'],
                               to_date_obj(r['echeance']), r['statut']])
                # Format dates colonnes A (Date) et F (Échéance)
                for col_idx in (1, 6):
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                             min_col=col_idx, max_col=col_idx):
                        for c in row:
                            c.number_format = 'DD/MM/YYYY'
                _autosize(ws)
                ws.freeze_panes = 'A2'

            buf = io.BytesIO()
            wb.save(buf)
            st.download_button(
                "📥 Télécharger (relances_commerciaux.xlsx)",
                data=buf.getvalue(),
                file_name=f"relances_commerciaux_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    with tab2:
        st.markdown("Dataset plat pour Power BI avec tranches d'âge et dernière relance.")
        if st.button("🔧 Générer l'export Power BI", type="primary"):
            pbi = df.copy()
            pbi['ecriture_date'] = pd.to_datetime(pbi['ecriture_date'],
                                                    errors='coerce',
                                                    dayfirst=True)
            pbi['annee'] = pbi['ecriture_date'].dt.year
            pbi['mois'] = pbi['ecriture_date'].dt.to_period('M').astype(str)
            pbi['age_jours'] = (pd.Timestamp.now().normalize() - pbi['ecriture_date']).dt.days

            def tranche(j):
                if pd.isna(j): return 'N/A'
                if j <= 30: return '0-30 j'
                if j <= 60: return '31-60 j'
                if j <= 90: return '61-90 j'
                if j <= 180: return '91-180 j'
                return '> 180 j'

            pbi['tranche_age'] = pbi['age_jours'].apply(tranche)
            pbi = pbi.merge(last_notes, on='comp_aux_num', how='left')

            out = io.BytesIO()
            with pd.ExcelWriter(out, engine='openpyxl') as wr:
                pbi.to_excel(wr, index=False, sheet_name='Creances')
                read_sheet('dossiers').to_excel(wr, index=False, sheet_name='Dossiers')
                read_sheet('notes').to_excel(wr, index=False, sheet_name='Notes')
            st.download_button(
                "📥 Télécharger (export_powerbi.xlsx)",
                data=out.getvalue(),
                file_name=f"creances_powerbi_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            st.caption("💡 Power BI : Obtenir les données → Excel → charger les 3 feuilles.")

    with tab3:
        st.markdown("Export des dossiers en contentieux groupés par responsable.")
        if df_ctx.empty:
            st.info("Aucun dossier en contentieux. Ajoutez-les dans "
                    "Import → Contentieux.")
        else:
            st.caption(f"{df_ctx['ref_client'].nunique()} dossier(s) — "
                       f"total {df_ctx['solde'].sum():,.0f} €".replace(",", " "))
            if st.button("🔧 Générer l'export contentieux", type="primary"):
                wb = openpyxl.Workbook()
                wb.remove(wb.active)

                # Feuille de synthèse
                ws = wb.create_sheet("Synthèse")
                headers = ['Responsable', 'Client', 'Ref dossier',
                           'Commercial', 'Solde (€)', 'Nb factures',
                           'Jours retard max', 'Dernière écriture']
                ws.append(headers)
                for c in ws[1]:
                    _style_header(c)

                synth_ct = df_ctx.groupby(
                    ['responsable', 'comp_aux_lib', 'ref_client',
                     'commercial'], dropna=False).agg(
                    solde=('solde', 'sum'),
                    nb=('piece_ref', 'count'),
                    jours=('jours_retard', 'max'),
                    derniere=('ecriture_date', 'max')
                ).reset_index().sort_values(['responsable', 'solde'],
                                             ascending=[True, False])

                for _, r in synth_ct.iterrows():
                    ws.append([r['responsable'], r['comp_aux_lib'], r['ref_client'],
                               r['commercial'],
                               round(r['solde'], 2), r['nb'],
                               int(r['jours']) if pd.notna(r['jours']) else '',
                               to_date_obj(r['derniere'])])

                total_row = ws.max_row + 1
                ws.cell(total_row, 1, 'TOTAL').font = Font(bold=True)
                ws.cell(total_row, 5, f'=SUM(E2:E{total_row - 1})').font = Font(bold=True)
                for row in ws.iter_rows(min_row=2, max_row=total_row,
                                         min_col=5, max_col=5):
                    for c in row:
                        c.number_format = '#,##0.00 €'
                # Format date colonne H (Dernière écriture)
                for row in ws.iter_rows(min_row=2, max_row=total_row - 1,
                                         min_col=8, max_col=8):
                    for c in row:
                        c.number_format = 'DD/MM/YYYY'
                _autosize(ws)
                ws.freeze_panes = 'A2'

                # Une feuille par responsable
                for resp in sorted(df_ctx['responsable'].dropna().unique()):
                    if not resp:
                        continue
                    df_r = df_ctx[df_ctx['responsable'] == resp] \
                        .sort_values(['comp_aux_lib', 'ecriture_date'])
                    safe = resp[:31].replace('/', '-').replace('\\', '-')
                    ws = wb.create_sheet(safe)
                    headers = ['Client', 'Ref dossier', 'Réf. pièce', 'Date',
                               'Journal', 'Libellé', 'Débit', 'Crédit', 'Solde',
                               'Jours retard', 'Commercial']
                    ws.append(headers)
                    for c in ws[1]:
                        _style_header(c)
                    for _, r in df_r.iterrows():
                        ws.append([r['comp_aux_lib'], r['ref_client'], r['piece_ref'],
                                   to_date_obj(r['ecriture_date']), r['journal_code'],
                                   r['ecriture_lib'],
                                   round(r['debit'], 2), round(r['credit'], 2),
                                   round(r['solde'], 2),
                                   int(r.get('jours_retard', 0) or 0),
                                   r['commercial']])
                    last = ws.max_row + 1
                    ws.cell(last, 1, 'TOTAL').font = Font(bold=True)
                    ws.cell(last, 9, f'=SUM(I2:I{last - 1})').font = Font(bold=True)
                    for row in ws.iter_rows(min_row=2, max_row=last,
                                             min_col=7, max_col=9):
                        for c in row:
                            c.number_format = '#,##0.00 €'
                    # Format date colonne D
                    for row in ws.iter_rows(min_row=2, max_row=last - 1,
                                             min_col=4, max_col=4):
                        for c in row:
                            c.number_format = 'DD/MM/YYYY'
                    _autosize(ws)
                    ws.freeze_panes = 'A2'

                buf = io.BytesIO()
                wb.save(buf)
                st.download_button(
                    "📥 Télécharger (export_contentieux.xlsx)",
                    data=buf.getvalue(),
                    file_name=f"contentieux_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

    with tab4:
        st.markdown("Synthèse pour la direction : un seul tableau avec les "
                    "résumés de note les plus récents.")
        if st.button("🔧 Générer l'export direction", type="primary"):
            wb = openpyxl.Workbook()
            wb.remove(wb.active)

            # Récupère le résumé direction (1 par client) depuis la feuille dédiée
            df_resumes = read_sheet('resumes')
            last_resume = pd.DataFrame(columns=['comp_aux_num', 'note_resume',
                                                 'action_resume',
                                                 'responsable_action',
                                                 'nature_creance',
                                                 'date_recouvrement',
                                                 'date_note', 'auteur'])
            if not df_resumes.empty and 'comp_aux_num' in df_resumes.columns:
                resumes_renamed = df_resumes.rename(columns={
                    'resume': 'note_resume',
                    'date_maj': 'date_note',
                })
                # Ajoute les colonnes manquantes si schéma ancien
                for col in ('action_resume', 'responsable_action',
                            'nature_creance', 'date_recouvrement'):
                    if col not in resumes_renamed.columns:
                        resumes_renamed[col] = ''
                last_resume = resumes_renamed[
                    ['comp_aux_num', 'note_resume', 'action_resume',
                     'responsable_action', 'nature_creance',
                     'date_recouvrement', 'date_note', 'auteur']
                ].drop_duplicates('comp_aux_num')

            # Synthèse globale (créances ouvertes y compris contentieux)
            df_dir = df_all.copy()

            # Exclusion des entités internes / intercos du rapport Direction
            EXCLUSIONS_DIRECTION = [
                'SCCV LHDL',
                'SCCV SAINT PAIR SANTE',
                'DPA DUD',
                'DESIGN PROMOTIONS',
                'SNC LA COUR AUX CHEVALIERS',
            ]
            if not df_dir.empty and 'comp_aux_lib' in df_dir.columns:
                lib_upper = df_dir['comp_aux_lib'].fillna('').astype(str).str.upper()
                mask_excl = pd.Series(False, index=df_dir.index)
                for excl in EXCLUSIONS_DIRECTION:
                    mask_excl |= lib_upper.str.contains(excl.upper(), na=False)
                df_dir = df_dir[~mask_excl]
                if mask_excl.any():
                    st.caption(f"ℹ️ {mask_excl.sum()} ligne(s) exclue(s) "
                               f"(entités internes : {', '.join(EXCLUSIONS_DIRECTION)})")

            ws = wb.create_sheet("Synthèse Direction")
            # Ordre :
            #  1=Client 2=Commercial 3=Conducteur 4=État 5=Avancement facture
            #  6=Nature créance 7=Date livraison 8=Date facture
            #  9=Date recouvrement estimée 10=Solde dû
            #  11=Consigné[H] 12=Solde net[H] 13=Jours retard[H]
            #  14=Prov risque[H] 15=Prov créances douteuses[H]
            #  16=Résumé 17=Action 18=Responsable
            #  19=Date MAJ 20=Auteur résumé 21=Statut
            headers = ['Client', 'Commercial', 'Conducteur', 'État',
                       'Avancement facture', 'Nature créance',
                       'Date livraison', 'Date facture',
                       'Date recouvrement estimée',
                       'Solde dû (€)', 'Consigné huissier (€)', 'Solde net (€)',
                       'Jours retard max',
                       'Prov. risque (€)', 'Prov. créances douteuses (€)',
                       'Résumé', 'Action à mener', 'Responsable action',
                       'Date MAJ', 'Auteur résumé', 'Statut']
            ws.append(headers)
            for c in ws[1]:
                _style_header(c)

            # Prépare les colonnes optionnelles
            df_dir = df_dir.copy()
            if 'situation' not in df_dir.columns:
                df_dir['situation'] = ''
            # Date de facture effective (cascade : PROGEMI/manuel > pièce > écriture)
            if 'date_facture_eff' not in df_dir.columns:
                df_dir['date_facture_eff'] = ''
            df_dir['date_facture_eff'] = df_dir['date_facture_eff'] \
                .fillna('').astype(str)
            df_dir['_dt_fact'] = pd.to_datetime(df_dir['date_facture_eff'],
                                                 errors='coerce')

            # Garde la situation et date facture de la facture la plus récente
            df_dir_sorted = df_dir.sort_values('_dt_fact', ascending=False)
            last_fact = df_dir_sorted.drop_duplicates('comp_aux_num')[
                ['comp_aux_num', 'situation', 'date_facture_eff']
            ].rename(columns={'situation': '_situation_last',
                               'date_facture_eff': '_date_fact_last'})

            synth_d = df_dir.groupby(['comp_aux_num', 'comp_aux_lib',
                                       'commercial', 'conducteur',
                                       'etat'], dropna=False).agg(
                solde=('solde', 'sum'),
                jours=('jours_retard', 'max'),
                contentieux=('contentieux', 'first'),
                provision_risque=('provision_risque', 'first'),
                provision_creances_douteuses=('provision_creances_douteuses',
                                                'first'),
                montant_consigne=('montant_consigne', 'first'),
                date_livraison=('date_reception', 'first'),
                date_plus_ancienne=('_dt_fact', 'min'),
            ).reset_index().sort_values('date_plus_ancienne',
                                          ascending=True,
                                          na_position='last')

            synth_d = synth_d.merge(last_fact, on='comp_aux_num', how='left')
            synth_d = synth_d.merge(last_resume, on='comp_aux_num', how='left')

            for _, r in synth_d.iterrows():
                statut = "⚖️ Contentieux" if r['contentieux'] else "Suivi commercial"
                consigne = float(r.get('montant_consigne', 0) or 0)
                solde_brut = float(r['solde'] or 0)
                solde_net = solde_brut - consigne
                ws.append([
                    r['comp_aux_lib'],
                    r['commercial'], r['conducteur'], r['etat'],
                    r.get('_situation_last', '') or '',
                    r.get('nature_creance', '') or '',
                    to_date_obj(r.get('date_livraison', '') or ''),
                    to_date_obj(r.get('_date_fact_last', '') or ''),
                    to_date_obj(r.get('date_recouvrement', '') or ''),
                    round(solde_brut, 2),
                    round(consigne, 2),
                    round(solde_net, 2),
                    int(r['jours']) if pd.notna(r['jours']) else '',
                    round(r['provision_risque'] or 0, 2),
                    round(r['provision_creances_douteuses'] or 0, 2),
                    r.get('note_resume', '') or '',
                    r.get('action_resume', '') or '',
                    r.get('responsable_action', '') or '',
                    to_date_obj(r.get('date_note', '') or ''),
                    r.get('auteur', '') or '',
                    statut,
                ])

            total_row = ws.max_row + 1
            ws.cell(total_row, 1, 'TOTAL').font = Font(bold=True)
            # Colonnes monétaires : 10=Solde, 11=Consigné, 12=Solde net,
            #                       14=Prov risque, 15=Prov créances douteuses
            for col_idx in (10, 11, 12, 14, 15):
                ws.cell(total_row, col_idx,
                        f'=SUM({get_column_letter(col_idx)}2:'
                        f'{get_column_letter(col_idx)}{total_row - 1})').font = Font(bold=True)
                for row in ws.iter_rows(min_row=2, max_row=total_row,
                                         min_col=col_idx, max_col=col_idx):
                    for c in row:
                        c.number_format = '#,##0.00 €'

            # Format dates : 7=Date livraison, 8=Date facture,
            #                9=Date recouvrement, 19=Date MAJ
            for col_idx in (7, 8, 9, 19):
                for row in ws.iter_rows(min_row=2, max_row=total_row - 1,
                                         min_col=col_idx, max_col=col_idx):
                    for c in row:
                        c.number_format = 'DD/MM/YYYY'

            _autosize(ws)

            # Masque : Consigné (K=11), Solde net (L=12), Jours retard (M=13),
            #          Prov. risque (N=14), Prov. créances douteuses (O=15)
            for letter in ('K', 'L', 'M', 'N', 'O'):
                ws.column_dimensions[letter].hidden = True

            # Fige les 5 premières colonnes + la ligne d'en-tête
            ws.freeze_panes = 'F2'

            # ============================================================
            # Onglet TCD : nature de créance × semaine d'encaissement prévue
            # ============================================================
            ws2 = wb.create_sheet("Encaissements par semaine")
            piv = synth_d.copy()
            # date_recouvrement est stocké en ISO (YYYY-MM-DD) → pas de dayfirst
            piv['_dt_recouv'] = pd.to_datetime(
                piv.get('date_recouvrement', ''), errors='coerce')
            piv = piv[piv['_dt_recouv'].notna()].copy()

            if piv.empty:
                ws2.append(["Aucune date de recouvrement estimée renseignée."])
                ws2.append(["Renseignez-les dans Notes & Relances → "
                            "Résumé direction."])
            else:
                piv['solde'] = pd.to_numeric(piv['solde'],
                                              errors='coerce').fillna(0)
                # Nature vide / placeholder → "Non renseigné"
                piv['nature_creance'] = piv['nature_creance'].fillna('') \
                    .astype(str).str.strip()
                piv.loc[piv['nature_creance'].isin(['', '—']),
                        'nature_creance'] = 'Non renseigné'
                # Lundi de la semaine = clé de tri ; libellé lisible
                piv['_lundi'] = (piv['_dt_recouv']
                                 - pd.to_timedelta(piv['_dt_recouv'].dt.weekday,
                                                    unit='D'))
                piv['_sem_label'] = (
                    'S' + piv['_dt_recouv'].dt.isocalendar().week
                    .astype(int).astype(str).str.zfill(2)
                    + ' (' + piv['_lundi'].dt.strftime('%d/%m/%Y') + ')')

                # Ordre chronologique des colonnes
                sem_order = (piv[['_lundi', '_sem_label']]
                             .drop_duplicates()
                             .sort_values('_lundi')['_sem_label'].tolist())

                pivot = pd.pivot_table(
                    piv, values='solde', index='nature_creance',
                    columns='_sem_label', aggfunc='sum', fill_value=0)
                pivot = pivot.reindex(columns=sem_order, fill_value=0)

                # En-tête
                header2 = ['Nature de créance'] + sem_order + ['Total']
                ws2.append(header2)
                for c in ws2[1]:
                    _style_header(c)

                for nature, prow in pivot.iterrows():
                    vals = [round(float(prow.get(s, 0) or 0), 2)
                            for s in sem_order]
                    ws2.append([nature] + vals + [round(sum(vals), 2)])

                # Ligne TOTAL par semaine
                tr2 = ws2.max_row + 1
                ws2.cell(tr2, 1, 'TOTAL').font = Font(bold=True)
                for j in range(2, len(header2) + 1):
                    cl = get_column_letter(j)
                    ws2.cell(tr2, j,
                             f'=SUM({cl}2:{cl}{tr2 - 1})').font = Font(bold=True)

                # Format € sur toutes les cellules de montant
                for row in ws2.iter_rows(min_row=2, max_row=tr2,
                                          min_col=2, max_col=len(header2)):
                    for c in row:
                        c.number_format = '#,##0.00 €'
                _autosize(ws2)
                ws2.freeze_panes = 'B2'

            # ============================================================
            # Onglet : Évolution du dû clients (avant / après 22/05/2026)
            # ============================================================
            ws3 = wb.create_sheet("Évolution du dû", 0)  # en 1er onglet

            CUTOFF_DU = pd.Timestamp(2026, 5, 22)
            AGE_SEUIL = 10  # jours
            # Référence figée des factures émises avant la coupure (snapshot 22/05/2026)
            FACTURES_AVANT_REF = 1111971.48

            raw = read_sheet('creances')
            # Date de référence = dernier import FEC
            ref_date = pd.Timestamp(datetime.now().date())
            if not raw.empty and 'import_date' in raw.columns:
                _imp = pd.to_datetime(raw['import_date'], errors='coerce')
                if _imp.notna().any():
                    ref_date = _imp.max().normalize()

            if raw.empty:
                ws3.append(["Aucune écriture FEC importée."])
            else:
                raw['debit'] = pd.to_numeric(raw['debit'],
                                              errors='coerce').fillna(0)
                raw['credit'] = pd.to_numeric(raw['credit'],
                                               errors='coerce').fillna(0)
                raw['_let'] = raw.get('ecriture_let', '').fillna('').astype(str) \
                    .str.strip()
                # En-cours = écritures non lettrées
                op = raw[raw['_let'] == ''].copy()
                op['_jrn'] = op.get('journal_code', '').fillna('') \
                    .astype(str).str.upper().str.strip()
                is_vt = op['_jrn'].str.startswith('VT')
                is_od = op['_jrn'].str.startswith('OD')

                # En-cours total = Total dû de l'application (même calcul que la sidebar)
                enr = load_creances_enrichies(only_open=True)
                enr = enr[enr['solde'].abs() > 0.01]
                en_cours = enr['solde'].sum()

                # Date d'émission effective par facture (PROGEMI > pièce > écriture)
                enr['_dem'] = pd.to_datetime(enr.get('date_facture_eff', ''),
                                              errors='coerce')
                enr['_dem'] = enr['_dem'].fillna(
                    pd.to_datetime(enr['ecriture_date'], errors='coerce'))

                # --- Bloc 2 : nouveau dû (factures émises APRÈS la coupure) ---
                new_df = enr[enr['_dem'] >= CUTOFF_DU].copy()
                new_total = new_df['solde'].sum()
                new_df['_age'] = (ref_date - new_df['_dem']).dt.days
                fact_sup10 = new_df[new_df['_age'] > AGE_SEUIL]['solde'].sum()
                fact_inf10 = new_df[new_df['_age'] <= AGE_SEUIL]['solde'].sum()

                # --- Bloc 1 : ancien dû (factures émises AVANT la coupure) ---
                # Baseline figée ; solde réel = en-cours - nouveau dû (réconcilie)
                deb_avant = FACTURES_AVANT_REF
                solde_old = en_cours - new_total
                # Mouvements depuis la coupure (crédits non lettrés par journal)
                avoirs = op[is_vt & (op['credit'] > 0)]['credit'].sum()
                compensation = op[is_od & (op['credit'] > 0)]['credit'].sum()
                paiements = op[(~is_vt) & (~is_od)
                               & (op['credit'] > 0)]['credit'].sum()
                # Écart pour réconcilier baseline - mouvements = solde réel
                autres = deb_avant - avoirs - paiements - compensation - solde_old

                ref_str = ref_date.strftime('%d/%m/%Y')
                cut_str = CUTOFF_DU.strftime('%d/%m/%Y')

                thin = Side(style='thin', color='2C3E50')
                box = Border(left=thin, right=thin, top=thin, bottom=thin)

                def _ligne(label, valeur, gras=False, encadre=False,
                           indent=False, neg=False):
                    r = ws3.max_row + 1
                    cl = ws3.cell(r, 1, ("   " if indent else "") + label)
                    cv = ws3.cell(r, 2, -valeur if neg else valeur)
                    cv.number_format = '#,##0.00 €'
                    if gras:
                        cl.font = Font(name='Segoe UI', bold=True, size=11)
                        cv.font = Font(name='Segoe UI', bold=True, size=11)
                    if encadre:
                        cl.fill = PatternFill('solid', start_color='EAF1E0')
                        cv.fill = PatternFill('solid', start_color='EAF1E0')
                        cl.border = box
                        cv.border = box
                    return r

                # Titre
                t = ws3.cell(1, 1, f"Évolution du dû clients — au {ref_str}")
                t.font = Font(name='Segoe UI', bold=True, size=14,
                              color='2C3E50')
                ws3.append([])

                # Bloc 1
                _ligne(f"Factures émises avant le {cut_str}", deb_avant,
                       gras=True, encadre=True)
                ws3.append([])
                _ligne("Avoirs émis", avoirs, indent=True, neg=True)
                _ligne("Paiements reçus", paiements, indent=True, neg=True)
                _ligne("Compensation", compensation, indent=True, neg=True)
                if abs(autres) > 0.01:
                    _ligne("Autres régularisations", autres, indent=True,
                           neg=True)
                ws3.append([])
                _ligne(f"Solde au {ref_str}", solde_old,
                       gras=True, encadre=True)
                ws3.append([])
                ws3.append([])

                # Bloc 2
                _ligne(f"Factures émises après le {cut_str}", new_total,
                       gras=True, encadre=True)
                _ligne(f"Émission factures > {AGE_SEUIL}j", fact_sup10,
                       indent=True)
                _ligne(f"Émission factures ≤ {AGE_SEUIL}j", fact_inf10,
                       indent=True)
                ws3.append([])
                ws3.append([])

                # Bloc 3
                _ligne(f"En cours clients au {ref_str}", en_cours,
                       gras=True, encadre=True)

                ws3.column_dimensions['A'].width = 38
                ws3.column_dimensions['B'].width = 18

            buf = io.BytesIO()
            wb.save(buf)
            st.download_button(
                "📥 Télécharger (export_direction.xlsx)",
                data=buf.getvalue(),
                file_name=f"direction_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )


# ============================================================
# PAGE ACCUEIL
# ============================================================
def page_accueil():
    user = current_user()
    if user is None:
        st.header("🏠 Bienvenue")
        st.info("Sélectionnez un utilisateur dans la barre latérale "
                "pour voir votre espace personnel.")
        st.markdown("**Aucun utilisateur actif** ? Ajoutez-en un dans "
                    "Import → onglet **Utilisateurs**.")
        return

    st.header(f"🏠 Bonjour {user['nom']}")
    st.caption(f"Connecté en tant que `{user['email']}`")

    # Récupère les notes de l'utilisateur (assignées ou créées par lui)
    notes = read_sheet('notes')
    if notes.empty:
        st.info("Aucune note enregistrée pour l'instant.")
        return

    if 'assigne_a' not in notes.columns:
        notes['assigne_a'] = ''
    notes['echeance'] = notes['echeance'].fillna('').astype(str)
    notes['statut'] = notes['statut'].fillna('Ouvert').astype(str)

    # Seules les notes explicitement assignées à l'utilisateur courant.
    # Si le champ assigne_a est vide ET que l'auteur est l'utilisateur, on retombe
    # sur les notes « auto-créées sans assignation » uniquement.
    user_email = user['email'].lower()
    mes_notes = notes[
        notes['assigne_a'].astype(str).str.lower() == user_email
    ].copy()
    mes_notes = mes_notes[mes_notes['statut'] != 'Résolu']

    if mes_notes.empty:
        st.success("✅ Aucune tâche en attente. Bonne journée !")
        return

    # Calcul de l'urgence par date d'échéance
    today = pd.Timestamp(datetime.now().date())
    mes_notes['_ech'] = pd.to_datetime(mes_notes['echeance'],
                                        errors='coerce', dayfirst=True)
    mes_notes['_jours_avant'] = (mes_notes['_ech'] - today).dt.days

    en_retard = mes_notes[mes_notes['_jours_avant'] < 0]
    urgent = mes_notes[(mes_notes['_jours_avant'] >= 0)
                        & (mes_notes['_jours_avant'] <= 7)]
    a_venir = mes_notes[mes_notes['_jours_avant'] > 7]
    sans_echeance = mes_notes[mes_notes['_ech'].isna()]

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("🔴 En retard", len(en_retard))
    c2.metric("🟠 Sous 7 jours", len(urgent))
    c3.metric("🟢 À venir", len(a_venir))
    c4.metric("⚪ Sans échéance", len(sans_echeance))

    def _show_section(title, df, color):
        if df.empty:
            return
        st.subheader(title)
        for _, n in df.sort_values('_ech').iterrows():
            with st.container(border=True):
                cc1, cc2, cc3 = st.columns([3, 2, 2])
                cc1.markdown(f"**Client** : `{n.get('comp_aux_num', '')}`")
                cc1.markdown(f"*{n.get('action', '') or '(note)'}*")
                cc1.write(n.get('note', ''))
                cc2.markdown(f"**Échéance** : {n.get('echeance', '') or '—'}")
                if pd.notna(n['_ech']):
                    j = int(n['_jours_avant'])
                    if j < 0:
                        cc2.markdown(f":red[En retard de {abs(j)} jour(s)]")
                    elif j == 0:
                        cc2.markdown(":orange[Échéance aujourd'hui]")
                    else:
                        cc2.markdown(f"Dans {j} jour(s)")
                cc3.markdown(f"**Statut** : {n.get('statut', '')}")
                cc3.caption(f"Créée le {n.get('date_note', '')}")

    _show_section("🔴 En retard", en_retard, "red")
    _show_section("🟠 À traiter sous 7 jours", urgent, "orange")
    _show_section("🟢 À venir", a_venir, "green")
    _show_section("⚪ Sans échéance", sans_echeance, "gray")


# ============================================================
# NAVIGATION
# ============================================================
PAGES = {
    "🏠 Accueil": page_accueil,
    "📥 Import": page_import,
    "📊 Créances": page_creances,
    "📝 Notes & Relances": page_notes,
    "📤 Export": page_export,
}

# Check config before any page
ok, msg = check_config()
if ok and not is_logged_in():
    show_login()
    st.stop()
if not ok:
    st.error(f"❌ Configuration manquante : {msg}")
    st.info("""
    **Pour configurer l'app localement :**

    1. Créez le fichier `.streamlit/secrets.toml` dans le dossier du projet
    2. Ajoutez le contenu de votre `google_credentials.json` au bon format
    3. Voir le fichier `.streamlit/secrets.toml.example` pour le modèle

    **Pour Streamlit Cloud :**
    Configurez les secrets dans l'interface (Settings → Secrets).
    """)
    st.stop()

with st.sidebar:
    st.title("💼 Suivi Créances")
    st.caption("DCA — Suivi clients")

    # --- Identification utilisateur ---
    _cur_user = current_user()
    if _cur_user is not None:
        st.markdown(f"👤 **{_cur_user['nom']}**")
        st.caption(_cur_user['email'])
        if st.button("🚪 Se déconnecter", use_container_width=True):
            logout_user()
            st.rerun()

    st.divider()
    page = st.radio("Navigation", list(PAGES.keys()), label_visibility="collapsed")
    st.divider()

    # Utilise les créances rapprochées (cohérent avec la page Créances)
    _df_side = load_creances_enrichies(only_open=True)
    if not _df_side.empty:
        _df_side = _df_side[_df_side['solde'].abs() > 0.01]
        st.metric("Clients en créance", _df_side['comp_aux_num'].nunique())
        st.metric("Total dû", f"{_df_side['solde'].sum():,.0f} €".replace(",", " "))

    # Date du dernier import FEC
    _df_raw = read_sheet('creances')
    if not _df_raw.empty and 'import_date' in _df_raw.columns:
        _import_dates = pd.to_datetime(_df_raw['import_date'],
                                        errors='coerce')
        _last_import = _import_dates.max()
        if pd.notna(_last_import):
            st.caption(f"📅 Dernier import FEC : "
                       f"{_last_import.strftime('%d/%m/%Y à %H:%M')}")

    if st.button("🔄 Rafraîchir les données"):
        clear_cache()
        st.rerun()

PAGES[page]()
